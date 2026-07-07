import streamlit as st
import pandas as pd
import numpy as np
import math
import re
import io
import os
import copy
import datetime
import gc
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="製造計画自動スケジュールシステム", page_icon="🚜", layout="wide")

st.title("🚜 製造計画全自動スケジュールシステム (カレンダー完全同期版)")
st.markdown("### エクセルを置くだけで、過去実績からスピードを学習し、超高速で指示書を出力します")

st.sidebar.markdown("## 🏢 工場の選択")
factory_mode = st.sidebar.selectbox("対象の工場を選択してください", ["本社", "関西工場"])

st.sidebar.markdown("---")
st.sidebar.markdown("## 📅 カレンダー・目標設定")

def _get_schedule_month(today):
    week_monday = today - datetime.timedelta(days=today.weekday())
    week_saturday = week_monday + datetime.timedelta(days=5)
    if today.month == 12:
        next_first = datetime.date(today.year + 1, 1, 1)
    else:
        next_first = datetime.date(today.year, today.month + 1, 1)
    if week_monday <= next_first <= week_saturday:
        return next_first.month
    return today.month

_current_schedule_month = _get_schedule_month(datetime.date.today())
_month_options = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
_default_month = f"{_current_schedule_month}月"
_default_idx = _month_options.index(_default_month) if _default_month in _month_options else 0
target_month = st.sidebar.selectbox("計画対象の月度を選択してください", _month_options, index=_default_idx)

default_start = datetime.date.today() + datetime.timedelta(days=1)
start_date = st.sidebar.date_input("🚜 製造スケジュール開始日", default_start)

HOLIDAYS_FILE = "holidays_local.csv"
# 実績レポートは工場別に独立して保存・維持する
JISSEKI_FILES = {"本社": "jisseki_honsha_local.csv", "関西工場": "jisseki_kansai_local.csv"}
_LEGACY_JISSEKI_FILE = "jisseki_local.csv"  # 旧形式（工場共通）互換用

def _jisseki_key(f_mode):
    return f"jisseki_data_{f_mode}"

def _load_jisseki(f_mode):
    """起動時にCSVから該当工場の実績データを復元してセッションに保持する"""
    key = _jisseki_key(f_mode)
    if key in st.session_state:
        return st.session_state[key]
    jf = JISSEKI_FILES.get(f_mode, _LEGACY_JISSEKI_FILE)
    paths = [jf]
    # 旧形式ファイルしかない場合の互換読み込み
    if not os.path.exists(jf) and os.path.exists(_LEGACY_JISSEKI_FILE):
        paths.append(_LEGACY_JISSEKI_FILE)
    for p in paths:
        if os.path.exists(p):
            for enc in ('utf-8', 'cp932'):
                try:
                    df_j = pd.read_csv(p, encoding=enc)
                    st.session_state[key] = df_j
                    return df_j
                except: pass
    return None

def _save_jisseki(df_j, f_mode):
    """該当工場の実績データをセッションとCSVの両方に保存する（他工場には影響しない）"""
    st.session_state[_jisseki_key(f_mode)] = df_j
    try:
        df_j.to_csv(JISSEKI_FILES.get(f_mode, _LEGACY_JISSEKI_FILE), index=False, encoding='utf-8')
    except: pass

def parse_jisseki(df_j):
    """実績レポートDFから品目コード→最多実績ラインの辞書を返す"""
    df_j = df_j.copy()
    df_j.columns = [str(c).strip() for c in df_j.columns]
    line_col = next((c for c in df_j.columns if '設備名' in c), None)
    code_col = next((c for c in df_j.columns if '品目コード' in c), None)
    qty_col  = next((c for c in df_j.columns if '製造良品数' in c
                     and '予算' not in c and '到達' not in c and '差異' not in c), None)
    result = {}
    if line_col and code_col and qty_col:
        df_j = df_j[[line_col, code_col, qty_col]].copy()
        df_j.columns = ['設備名', '品目コード', '良品数']
        df_j['品目コード'] = df_j['品目コード'].astype(str).str.strip()
        df_j['良品数']     = pd.to_numeric(df_j['良品数'], errors='coerce').fillna(0)
        df_j['設備名']     = df_j['設備名'].ffill()
        df_j = df_j[~df_j['設備名'].astype(str).isin({'(なし)', 'nan', ''})]
        df_j = df_j[df_j['品目コード'].str.match(r'^[KH]\d')]  # 関西=K・本社=H の両方に対応
        df_j['ライン'] = df_j['設備名'].astype(str).str.extract(r'(\d+号機)')
        df_j = df_j.dropna(subset=['ライン'])
        grp = df_j.groupby(['品目コード', 'ライン'])['良品数'].sum().reset_index()
        for code_j, grp_df in grp.groupby('品目コード'):
            best_line = grp_df.loc[grp_df['良品数'].idxmax(), 'ライン']
            result[code_j] = best_line
    return result

# =====================================================================
# 🌟 実績ライン学習辞書（品目コード→ライン）の永続化
#   Streamlit Cloudではアプリ再起動（デプロイ・スリープ）でローカルCSVが
#   消えるため、①累積マージで学習を蓄積、②辞書CSVをダウンロードして
#   GitHubリポジトリに同梱すれば再起動後も恒久保持できる仕組みとする。
# =====================================================================
JISSEKI_DICT_FILES = {"本社": "jisseki_line_honsha.csv", "関西工場": "jisseki_line_kansai.csv"}

def _load_line_dict(f_mode):
    """学習済みライン辞書を読み込む（セッション→保存CSV→旧形式rawの順）"""
    key = f"line_dict_{f_mode}"
    if key in st.session_state:
        return st.session_state[key]
    d = {}
    p = JISSEKI_DICT_FILES.get(f_mode)
    if p and os.path.exists(p):
        for enc in ('utf-8', 'cp932'):
            try:
                dfd = pd.read_csv(p, encoding=enc)
                d = dict(zip(dfd['品目コード'].astype(str).str.strip(),
                             dfd['ライン'].astype(str).str.strip()))
                break
            except: pass
    if not d:
        # 旧形式（生データCSV）からの移行
        _raw = _load_jisseki(f_mode)
        if _raw is not None:
            try:
                d = parse_jisseki(_raw)
                _pfx_mig = 'H' if f_mode == "本社" else 'K'
                d = {k: v for k, v in d.items() if str(k).startswith(_pfx_mig)}
            except: pass
    st.session_state[key] = d
    return d

def _save_line_dict(d, f_mode):
    """学習辞書をセッションとCSVの両方へ保存"""
    st.session_state[f"line_dict_{f_mode}"] = d
    try:
        pd.DataFrame({'品目コード': list(d.keys()), 'ライン': list(d.values())}).to_csv(
            JISSEKI_DICT_FILES.get(f_mode), index=False, encoding='utf-8')
    except: pass

# 起動時に選択中工場の学習辞書を復元
_load_line_dict(factory_mode)

def _load_holidays():
    if 'holidays_data' in st.session_state:
        return st.session_state['holidays_data']
    if os.path.exists(HOLIDAYS_FILE):
        try:
            df_h = pd.read_csv(HOLIDAYS_FILE, encoding='utf-8')
            dates = [datetime.datetime.strptime(d, "%Y-%m-%d").date() for d in df_h['date'].tolist()]
            st.session_state['holidays_data'] = dates
            return dates
        except: pass
    return []

def _save_holidays(dates):
    st.session_state['holidays_data'] = dates
    try:
        df_h = pd.DataFrame({'date': [d.strftime("%Y-%m-%d") for d in dates]})
        df_h.to_csv(HOLIDAYS_FILE, index=False, encoding='utf-8')
    except: pass

_saved_holidays = _load_holidays()

st.sidebar.markdown("### 🛑 工場休業日の登録")
st.sidebar.caption("カレンダーの日付をタップするたびに 登録⇔解除 が切り替わります（赤=休業日）。土日は自動スキップのため登録不要です。")

import calendar as _cal_mod

# 表示月のセッション管理
if 'hol_cal_month' not in st.session_state:
    st.session_state['hol_cal_month'] = datetime.date(start_date.year, start_date.month, 1)

_nav1, _nav2, _nav3 = st.sidebar.columns([1, 2, 1])
with _nav1:
    if st.button("◀", key="hol_prev", use_container_width=True):
        _cm = st.session_state['hol_cal_month']
        st.session_state['hol_cal_month'] = (_cm - datetime.timedelta(days=1)).replace(day=1)
        st.rerun()
with _nav2:
    _cm_disp = st.session_state['hol_cal_month']
    st.markdown(f"<div style='text-align:center;font-weight:bold;padding-top:6px;'>{_cm_disp.year}年 {_cm_disp.month}月</div>", unsafe_allow_html=True)
with _nav3:
    if st.button("▶", key="hol_next", use_container_width=True):
        _cm = st.session_state['hol_cal_month']
        _nm_y = _cm.year + (1 if _cm.month == 12 else 0)
        _nm_m = 1 if _cm.month == 12 else _cm.month + 1
        st.session_state['hol_cal_month'] = datetime.date(_nm_y, _nm_m, 1)
        st.rerun()

_cm = st.session_state['hol_cal_month']
_weeks = _cal_mod.Calendar(firstweekday=0).monthdatescalendar(_cm.year, _cm.month)

# 曜日ヘッダー
_hdr_cols = st.sidebar.columns(7)
for _i, _wd in enumerate(["月", "火", "水", "木", "金", "土", "日"]):
    _hdr_cols[_i].markdown(f"<div style='text-align:center;font-size:11px;color:#888;'>{_wd}</div>", unsafe_allow_html=True)

# 日付グリッド（タップでトグル）
_hol_set = set(_saved_holidays)
for _week in _weeks:
    _day_cols = st.sidebar.columns(7)
    for _i, _d in enumerate(_week):
        if _d.month != _cm.month:
            _day_cols[_i].markdown("&nbsp;", unsafe_allow_html=True)
            continue
        if _d.weekday() >= 5:
            _day_cols[_i].markdown(f"<div style='text-align:center;color:#ccc;font-size:13px;padding:6px 0;'>{_d.day}</div>", unsafe_allow_html=True)
            continue
        _is_hol = _d in _hol_set
        if _day_cols[_i].button(
            str(_d.day),
            key=f"hol_{_d.isoformat()}",
            use_container_width=True,
            type="primary" if _is_hol else "secondary"
        ):
            if _is_hol:
                _hol_set.discard(_d)
            else:
                _hol_set.add(_d)
            _save_holidays(sorted(_hol_set))
            st.rerun()

# 登録済み一覧と一括削除
if _saved_holidays:
    _w_kanji_list = ["月", "火", "水", "木", "金", "土", "日"]
    _list_str = "、".join(f"{d.month}/{d.day}({_w_kanji_list[d.weekday()]})" for d in sorted(_saved_holidays))
    st.sidebar.caption(f"📋 登録済み {len(_saved_holidays)}日: {_list_str}")
    if st.sidebar.button("🗑️ 全て削除", key="hol_clear"):
        _save_holidays([])
        st.rerun()

holidays_input = list(_saved_holidays)

def _get_month_end(month_label, base_today):
    m = int(month_label.replace("月", ""))
    y = base_today.year
    if m < base_today.month - 6:
        y += 1
    if m == 12:
        return datetime.date(y, 12, 31)
    else:
        return datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)

plan_to_yearend = st.sidebar.checkbox(
    "📆 期末（10月末）まで一括計画する",
    value=False,
    help="ONにすると、選択月度から10月度（期末）までの全月の計画残数を合算して、10月末までのスケジュールを一括生成します。"
)

def _get_yearend_date(base_today):
    """当期の期末（10月31日）を返す。11月・12月なら翌年の10月末。"""
    y = base_today.year if base_today.month <= 10 else base_today.year + 1
    return datetime.date(y, 10, 31)

if plan_to_yearend:
    _month_end = _get_yearend_date(datetime.date.today())
else:
    _month_end = _get_month_end(target_month, datetime.date.today())

def _count_business_days(s_date, e_date, holiday_list):
    if e_date < s_date:
        return 1
    cnt = 0
    d = s_date
    while d <= e_date:
        if d.weekday() < 5 and d not in holiday_list:
            cnt += 1
        d += datetime.timedelta(days=1)
    return max(cnt, 1)

_auto_target_days = _count_business_days(start_date, _month_end, holidays_input)

_max_days = 260 if plan_to_yearend else 31
target_days = st.sidebar.number_input(
    "目標稼働日数 (この日数以内に作り切る)" if plan_to_yearend else "当月の目標稼働日数 (この日数以内に作り切る)",
    min_value=1, max_value=_max_days,
    value=min(_auto_target_days, _max_days)
)
st.sidebar.caption(f"📌 {start_date.strftime('%Y/%m/%d')}〜{_month_end.strftime('%Y/%m/%d')}の営業日数を自動計算：{_auto_target_days}日")

st.sidebar.markdown("---")
st.sidebar.markdown("## 1. ファイルのアップロード")
file_zai = st.sidebar.file_uploader("① 在庫推移リスト (Excel形式: .xlsx)", type=["xlsx"])

if factory_mode == "本社":
    file_gekkan = st.sidebar.file_uploader("② 本社 月間製造計画書 (Excel形式: .xlsx)", type=["xlsx"])
else:
    file_gekkan = st.sidebar.file_uploader("② 関西工場 月間製造計画書 (Excel形式: .xlsx)", type=["xlsx"])

file_bom = st.sidebar.file_uploader("③ [任意] 新しいBOM構成表マスタ (ExcelまたはCSV)", type=["xlsx", "csv"])
file_jisseki = st.sidebar.file_uploader("④ [任意] 製造実績レポート (Excel形式: .xlsx)", type=["xlsx"])
file_itemmaster = st.sidebar.file_uploader(
    "⑤ [任意] 品目マスタ (袋サイズ切り替えルール用・CSV)", type=["csv", "xlsx"],
    help="「幅×ピッチ（㎜）」列を使い、同じ容量（例:25L）でも実際の袋型枠サイズが近い品目を連続製造するよう切り替え順序を最適化します。"
)

st.sidebar.markdown("---")
consider_iko = st.sidebar.checkbox(
    "📦 確定済みの入庫予定（入）を差し引いて不足数を計算する",
    value=True,
    help="ONにすると、在庫推移リストの「入」行に記載済みの入庫予定数量を、製造計画期間の安全割れ不足数からあらかじめ差し引きます。さらに、その確定済み入庫予定はシート2・3にも「製造指示済」として反映され、その時間分を差し引いた残り時間で新規の製造計画が組まれます。OFFの場合は今日時点の在庫のみで判定します。"
)

if factory_mode == "本社":
    rule_info = "・定時時間: 月〜木 430分(16:30終) / 金曜 400分(16:00終・メンテ)\n・稼働ライン: 2号機、3号機、5号機、6号機、その他"
else:
    rule_info = "・定時時間: 月〜木 430分(16:30終) / 金曜 400分(16:00終・メンテ)\n・稼働ライン: 1号, 2号, 3号, 4号, 5号, 6号, その他"

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚙️ 現場同期・固定ルール")
st.sidebar.info(
    f"・選択中の工場: {factory_mode}\n"
    f"・対象月度: {target_month}度計画\n"
    f"・開始日: {start_date.strftime('%Y/%m/%d')}\n"
    f"{rule_info}\n"
    "・完全自動化: 人間による手動データ加工を一切排除した現場直結仕様\n"
    "・残業最適化: 労務管理優先、必ず30分刻みジャストで終了探索\n"
    "・製造理由: [現在庫がマイナス] [安全在庫割れ] [計画未達] [製造指示済] の4種仕分け\n"
    "・休憩ロック: 10:00(10分), 12:00(60分), 15:00(10分)"
)

# =====================================================================
# 🌟 最上流グローバル・独立パーサー関数群
# =====================================================================

def safe_seek(f):
    if hasattr(f, 'seek'): f.seek(0)

def fetch_github_file_bytes(path):
    """MRPアプリ（発注リスケ提案ツール）と共有しているGitHubリポジトリからファイルを取得する。
    st.secrets["github"]["token"]/["repo"] は mrp_link.py の自動連携と同じ設定を流用する。"""
    gh = st.secrets.get("github", {})
    token, repo = gh.get("token"), gh.get("repo")
    if not token or not repo:
        return None
    try:
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github.raw",
            "X-GitHub-Api-Version": "2022-11-28",
        }
        url = f"https://api.github.com/repos/{repo}/contents/{path}"
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code == 200:
            return r.content
    except Exception:
        pass
    return None

def read_item_master_bytes(raw_bytes, filename="ItemMaster.csv"):
    """品目マスタのバイト列をDataFrameに変換する。MRPアプリのloaders/item_master.pyと同じ
    想定フォーマット（CP932想定・cp932→utf-8-sig→utf-8の順で試行）に合わせる。"""
    if filename.lower().endswith(('.xlsx', '.xls')):
        return pd.read_excel(io.BytesIO(raw_bytes))
    for enc in ('cp932', 'utf-8-sig', 'utf-8'):
        try:
            return pd.read_csv(io.BytesIO(raw_bytes), encoding=enc)
        except Exception:
            continue
    return None

def load_excel_sheets_merged(file, keywords, exclude_keywords=None):
    safe_seek(file)
    xl = pd.ExcelFile(file)
    matched_sheets = [sheet for sheet in xl.sheet_names if any(kw in sheet for kw in keywords)]
    if exclude_keywords:
        matched_sheets = [sheet for sheet in matched_sheets if not any(ex in sheet for ex in exclude_keywords)]
    load_excel_sheets_merged.last_matched = list(matched_sheets)
    if not matched_sheets:
        load_excel_sheets_merged.last_matched = [f"{xl.sheet_names[0]}（キーワード不一致のため先頭シートを使用）"]
        safe_seek(file)
        df_single = pd.read_excel(file, sheet_name=0, header=None)
        return df_single

    base_df = pd.read_excel(xl, sheet_name=matched_sheets[0], header=None)
    item_row_idx = 1
    for i in range(min(15, len(base_df))):
        row_vals = [str(v).strip() for v in base_df.iloc[i].values]
        if any(k in row_vals for k in ['品目コード', '品目ｺｰﾄﾞ', '商品コード', '商品CD', '商品CODE']):
            item_row_idx = i; break

    for sheet in matched_sheets[1:]:
        add_df = pd.read_excel(xl, sheet_name=sheet, header=None)
        if len(add_df) > item_row_idx + 1:
            tmp_df = add_df.iloc[item_row_idx + 1:].copy()
            del add_df
            base_df = pd.concat([base_df, tmp_df], ignore_index=True)
            del tmp_df
            gc.collect()

    del xl
    gc.collect()
    return base_df

def clean_bom_master(df_raw_bom):
    if df_raw_bom is None or df_raw_bom.empty: return None
    h_row = 0
    for i in range(min(15, len(df_raw_bom))):
        row_vals = [str(v).strip() for v in df_raw_bom.iloc[i].values]
        if any(k in row_vals for k in ['品目コード', '商品コード', '商品CODE', '配合CODE', '配合コード', '親品目コード']):
            h_row = i; break
    df_clean = df_raw_bom.iloc[h_row+1:].copy()
    df_clean.columns = [str(c).strip() for c in df_raw_bom.iloc[h_row].values]
    return df_clean

def extract_volume_safe(name_str):
    n_str = str(name_str)
    if '真砂土' in n_str and '15' in n_str:
        return 12
    kg_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:[kKｋＫ][gGｇＧ])', n_str)
    if kg_match:
        try: return -float(kg_match.group(1))
        except: return 14
    l_match = re.search(r'(\d+(?:\.\d+)?)\s*[LLｌｌＬＬ]', n_str)
    if l_match:
        try: return float(l_match.group(1))
        except: return 14
    elif '特大袋' in n_str: return 55
    g_match = re.search(r'(\d+(?:\.\d+)?)\s*[gGｇＧ](?![gGｇＧ])', n_str)
    if g_match and not re.search(r'[kKｋＫ][gGｇＧ]', n_str):
        try: return -float(g_match.group(1)) / 1000.0
        except: return 14
    else: return 14

def is_kg_product(name_str):
    n_str = str(name_str)
    if re.search(r'\d+(?:\.\d+)?\s*[kKｋＫ][gGｇＧ]', n_str):
        return True
    if re.search(r'\d+(?:\.\d+)?\s*[gGｇＧ](?![gGｇＧ])', n_str) and not re.search(r'[kKｋＫ][gGｇＧ]', n_str):
        return True
    return False

def get_kg_weight(name_str):
    n_str = str(name_str)
    kg_match = re.search(r'(\d+(?:\.\d+)?)\s*[kKｋＫ][gGｇＧ]', n_str)
    if kg_match:
        return float(kg_match.group(1))
    g_match = re.search(r'(\d+(?:\.\d+)?)\s*[gGｇＧ]', n_str)
    if g_match:
        return float(g_match.group(1)) / 1000
    return 1.0

def extract_core_name(name_str):
    n_str = str(name_str)
    n_str = re.sub(r'[（(][^）)]*[）)]', '', n_str)
    n_str = re.sub(r'\d+(?:\.\d+)?\s*[LLｌｌＬＬkKｋＫgGｇＧ]+', '', n_str)
    for size_word in ['細粒', '中粒', '大粒', '小粒', '特大袋', '特大', 'ミニ', '大', '中', '小']:
        n_str = n_str.replace(size_word, '')
    n_str = re.sub(r'[・･\s　/／\-]', '', n_str)
    n_str = re.sub(r'^(PB|new|New|NEW)', '', n_str)
    return n_str.strip()

def is_similar_product(core_a, core_b):
    if not core_a or not core_b:
        return False
    return core_a == core_b or core_a in core_b or core_b in core_a

def parse_bag_size(val):
    """品目マスタの「幅×ピッチ（㎜）」列を (幅, ピッチ, 追加値) のタプルに正規化する。
    同じ容量_L（例:25L）でも実際の袋型枠は微妙に異なる（例:450x650 と 460x650）ため、
    容量とは別に実際の袋サイズで切り替え順序を判定するのに使う。"""
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    s = s.translate(str.maketrans({'ｘ': 'x', 'Ｘ': 'x', '×': 'x', 'X': 'x'}))
    parts = s.split('x')
    if len(parts) < 2:
        return None
    def _num(token):
        m = re.search(r'\d+(?:\.\d+)?', token)
        return float(m.group()) if m else None
    w = _num(parts[0])
    p_token = parts[1]
    extra = 0.0
    if '+' in p_token:
        p_str, extra_str = p_token.split('+', 1)
        p = _num(p_str)
        extra = _num(extra_str) or 0.0
    else:
        p = _num(p_token)
    if w is None or p is None:
        return None
    return (w, p, extra)

def bag_size_diff(a, b):
    if not isinstance(a, tuple) or not isinstance(b, tuple):
        return None
    return sum(abs(x - y) for x, y in zip(a, b))

def sort_jobs_by_size_proximity(df_line):
    unprocessed = df_line.to_dict('records')
    if not unprocessed: return []
    for j in unprocessed:
        if 'コア名称' not in j:
            j['コア名称'] = extract_core_name(j.get('品目名', ''))
    processed = []
    first_recipe = unprocessed[0]['中身設計コード']
    same_recipe_jobs = [j for j in unprocessed if j['中身設計コード'] == first_recipe]
    same_recipe_jobs.sort(key=lambda x: x['容量_L'], reverse=True)
    processed.extend(same_recipe_jobs)
    for j in same_recipe_jobs: unprocessed.remove(j)
    while unprocessed:
        last_job = processed[-1]
        last_vol = last_job['容量_L']
        last_core = last_job['コア名称']
        last_bag = last_job.get('袋サイズ')
        if not isinstance(last_bag, tuple):
            last_bag = None
        # 袋サイズ（幅×ピッチ）が完全一致する品目を最優先で連続製造し、資材（フィルム）交換を減らす
        same_bag_candidates = [j for j in unprocessed if last_bag and isinstance(j.get('袋サイズ'), tuple) and j.get('袋サイズ') == last_bag]
        similar_candidates = [j for j in unprocessed if is_similar_product(last_core, j['コア名称'])]
        if same_bag_candidates:
            same_bag_candidates.sort(key=lambda x: (abs(x['容量_L'] - last_vol), x['グループ緊急度']))
            best_idx = unprocessed.index(same_bag_candidates[0])
        elif similar_candidates:
            similar_candidates.sort(key=lambda x: (abs(x['容量_L'] - last_vol), x['グループ緊急度']))
            best_idx = unprocessed.index(similar_candidates[0])
        else:
            def _rank(j):
                bd = bag_size_diff(last_bag, j.get('袋サイズ'))
                return (bd if bd is not None else float('inf'), abs(j['容量_L'] - last_vol), j['グループ緊急度'])
            best_idx = min(range(len(unprocessed)), key=lambda idx: _rank(unprocessed[idx]))
        next_recipe = unprocessed[best_idx]['中身設計コード']
        same_recipe_jobs = [j for j in unprocessed if j['中身設計コード'] == next_recipe]
        same_recipe_jobs.sort(key=lambda x: x['容量_L'], reverse=True)
        processed.extend(same_recipe_jobs)
        for j in same_recipe_jobs: unprocessed.remove(j)
    return processed

def job_can_support(l_key, job_item, f_mode):
    """ライン l_key がジョブを応援製造できるか判定。
    get_capable_linesを真の物理制約として使用し、固定コードの誤応援を防ぐ。"""
    # 物理的に対応可能かチェック（固定コード・サイズ制約を一元管理）
    capable = get_capable_lines(
        job_item.get('容量_L', 0),
        job_item.get('品目名', ''),
        job_item.get('品目コード', ''),
        job_item.get('堆肥・腐葉土フラグ', False),
        f_mode
    )
    if l_key not in capable:
        return False

    # 追加制約（本社）
    if f_mode == "本社":
        if str(job_item.get('品目コード', '')) in ('H0690020', 'H0690000', 'H0690030', 'H0390000'):
            return False
        return not job_item.get('堆肥・腐葉土フラグ', False)

    # 追加制約（関西工場）
    else:
        if str(job_item.get('品目コード', '')).startswith('K0225'):
            return False
        name = str(job_item.get('品目名', ''))
        if any(k in name for k in _KEYWORDS_4GO):
            return False
        if '化成肥料' in name and 'ｺｰﾅﾝ' in name:
            return False
        vol = job_item.get('容量_L', 0)
        if vol < 0:
            return False
        if l_key == '4号機' and job_item.get('製造理由', '') != '現在庫がマイナス':
            return False
        if job_item.get('堆肥・腐葉土フラグ', False) or any(k in name for k in ['再生材', 'もう一土元気']):
            return False
        return True

def get_next_w_date(cur, holidays_list):
    nd = cur
    while nd.weekday() >= 5 or nd in holidays_list: nd += datetime.timedelta(days=1)
    return nd

SPEED_4GO = {
    'K0270430': 214,
    'K0521190': 51,
    'K0571080': 140,
    'K0670290': 336,
    'K0101700': 187,
    'K0101800': 182,
    'K0130200': 96,
    'K0130300': 100,
    'K0130660': 73,
    'K0190010': 54,
    'K0190900': 30,
    'K0400000': 197,
    'K0400010': 540,
    'K0400020': 212,
    'K0425010': 86,
    'K0430110': 50,
    'K0430120': 70,
    'K0430130': 55,
    'K0430140': 33,
    'K0465170': 178,
    'K0465900': 240,
    'K0466000': 175,
    'K0466100': 100,
    'K0480090': 51,
    'K0490040': 18,
    'K0490050': 38,
    'K0490080': 35,
}
SPEED_4GO_DEFAULT = 100

# =====================================================================
# 🌟 ライン物理対応・同一配合統一ロジック
# =====================================================================

# 速度優先度（袋/時・高いほど優先）
LINE_SPEED_PRIORITY = {
    "本社":     {'5号機': 650, '2号機': 400, '6号機': 260, '3号機': 150, 'その他': 0},
    "関西工場": {'5号機': 646, '6号機': 597, '2号機': 494, '1号機': 388, '3号機': 150, '4号機': 100, 'その他': 0}
}

# キーワードをグローバルで共有（本社・関西両方で使用）
_KEYWORDS_4GO = ['ピートモス', 'くん炭', 'バーミキュライト', 'パーライト',
                 'ﾋﾟｰﾄﾓｽ', 'ﾊﾞｰﾐｷｭﾗｲﾄ', 'ﾊﾟｰﾗｲﾄ']

# 特殊清掃が必要な品目キーワード（各稼働日の最初に製造することで、清掃を1日1回で済ませる）
SPECIAL_CLEANING_KEYWORDS = ['種まき培土', 'ピートモス', 'オーガニック', '軽石', '鉢底石',
                              'ﾋﾟｰﾄﾓｽ', 'ｵｰｶﾞﾆｯｸ']

# 本社ライン固定商品一覧（ライン固定一覧20260706.xlsxのD列「指定機械」より・54品目）
HONSHA_FIXED_LINE = {
    'H0120620': '2号機', 'H0120630': '2号機', 'H0220230': '2号機', 'H0500110': '2号機',
    'H0500120': '2号機', 'H0510140': '2号機', 'H0520870': '2号機', 'H0521060': '2号機',
    'H0100600': '3号機', 'H0101800': '3号機', 'H0190000': '3号機', 'H0190100': '3号機',
    'H0191000': '3号機', 'H0200050': '3号機', 'H0280520': '3号機', 'H0280560': '3号機',
    'H0280570': '3号機', 'H0280580': '3号機', 'H0280590': '3号機', 'H0300000': '3号機',
    'H0300080': '3号機', 'H0400000': '3号機', 'H0400020': '3号機', 'H0400400': '3号機',
    'H0420650': '3号機', 'H0420750': '3号機', 'H0480050': '3号機', 'H0480120': '3号機',
    'H0480140': '3号機', 'H0520820': '3号機', 'H0581140': '3号機', 'H0581150': '3号機',
    'H0680170': '3号機',
    'H0101210': '5号機', 'H0101220': '5号機', 'H0101230': '5号機', 'H0101240': '5号機',
    'H0225000': '5号機', 'H0225270': '5号機', 'H0225300': '5号機', 'H0410620': '5号機',
    'H0581260': '5号機',
    'H0290040': '6号機', 'H0290130': '6号機', 'H0290150': '6号機', 'H0390080': '6号機',
    'H0581190': '6号機', 'H0590060': '6号機', 'H0590190': '6号機', 'H0590250': '6号機',
    'H0590440': '6号機', 'H0590450': '6号機', 'H0590670': '6号機', 'H0590800': '6号機',
}

def get_capable_lines(vol, name, code, is_compost, f_mode):
    """物理的に製造可能なラインのリストを返す（速度優先ではなく物理制約のみ）"""
    code_str = str(code)
    name_str = str(name)

    if f_mode == "本社":
        # 固定コード
        if code_str in ('H0690020', 'H0690000', 'H0690030', 'H0390000'):
            return ['6号機']
        # ライン固定商品一覧（20260706）に記載の品目
        if code_str in HONSHA_FIXED_LINE:
            return [HONSHA_FIXED_LINE[code_str]]
        # 堆肥・腐葉土・再生材系
        if code_str == 'H0620030' or any(k in name_str for k in ['再生材', 'もう一土元気']) or is_compost:
            return ['3号機']
        # 9L以下は外注
        if vol < 10:
            return ['その他']
        capable = []
        if 10 <= vol <= 25:  capable.append('5号機')   # 5号機：10〜25L
        if 14 <= vol <= 26:  capable.append('2号機')   # 2号機：14〜26L
        if vol >= 25:        capable.append('6号機')   # 6号機：25L以上
        return capable if capable else ['その他']

    else:  # 関西工場
        if any(k in name_str for k in _KEYWORDS_4GO):
            return ['4号機']
        if code_str in ('K0390110', 'K0480080', 'K0680190'):
            return ['3号機']
        if code_str in ('K0270450', 'K0190010'):
            return ['その他']
        if code_str == 'K0430120':
            return ['4号機']
        if 'CLEAR' in name_str and 'ERA' in name_str:
            return ['その他']
        if '有機石灰' in name_str:
            return ['その他']
        if is_kg_product(name_str) and get_kg_weight(name_str) < 1.0:
            return ['4号機']
        is_compost_flag = any(k in name_str for k in ['腐葉土', '堆肥'])
        is_special = any(k in name_str for k in ['再生材', 'もう一土元気'])
        if is_compost or is_special:
            if '特大袋' in name_str:   return ['3号機']
            elif vol >= 40:            return ['1号機']
            elif vol >= 14:            return ['3号機']
            elif is_compost_flag:      return ['4号機']
            elif vol >= 1.2:           return ['5号機']
            else:                      return ['その他']
        if '化成肥料' in name_str and 'ｺｰﾅﾝ' in name_str:
            return ['5号機']
        if code_str == 'K0630390':
            return ['5号機']
        # kg品：kg重量をL換算の代理値として使用
        if vol < 0:
            kg_w = get_kg_weight(name_str)
            if kg_w < 1.0:
                return ['4号機']
            eff_vol = int(kg_w)
            capable = []
            if eff_vol < 10:           capable.append('5号機')
            if 10 <= eff_vol <= 20:    capable.append('6号機')  # 6号機：10〜20L
            if 10 <= eff_vol <= 25:    capable.append('2号機')  # 2号機：10〜25L
            if eff_vol >= 25:          capable.append('1号機')  # 1号機：25L以上
            return capable if capable else ['その他']
        # 通常品
        if vol < 1.2:
            return ['4号機']
        capable = []
        if 1.2 <= vol < 10:   capable.append('5号機')
        if 10 <= vol <= 20:   capable.append('6号機')  # 6号機：10〜20L
        if 10 <= vol <= 25:   capable.append('2号機')  # 2号機：10〜25L
        if vol >= 25:         capable.append('1号機')  # 1号機：25L以上
        return capable if capable else ['その他']


def unify_recipe_lines(df, f_mode):
    """同一配合の製品を可能な限り同一ラインに統一する。
    全製品に共通して物理対応可能なラインがある場合のみ統一し、
    物理的に不可能な場合は個別割り当てを維持する。"""
    speed_priority = LINE_SPEED_PRIORITY[f_mode]
    df = df.copy()

    for recipe_code, group in df.groupby('中身設計コード'):
        lines_used = set(group['製造ライン'].unique())
        if len(lines_used) == 1:
            continue  # 既に同一ライン → スキップ

        # 各製品の物理的に対応可能なラインセットを収集
        capable_sets = []
        for _, row in group.iterrows():
            capable = get_capable_lines(
                row['容量_L'], row['品目名'], row['品目コード'],
                row['堆肥・腐葉土フラグ'], f_mode
            )
            capable_sets.append(set(capable))

        # 全製品で共通対応可能なラインの積集合
        common_lines = capable_sets[0]
        for s in capable_sets[1:]:
            common_lines = common_lines & s

        if not common_lines:
            continue  # 物理的に共通ラインなし → 分散製造を維持

        # 最速の共通ラインに統一
        best_line = max(common_lines, key=lambda l: speed_priority.get(l, 0))
        df.loc[df['中身設計コード'] == recipe_code, '製造ライン'] = best_line

    return df


def get_sp(line, vol, f_mode, item_code=''):
    if f_mode == "関西工場" and line == '5号機' and str(item_code).startswith('K0225') and vol == 12:
        return 490
    if f_mode == "関西工場" and line == '5号機' and str(item_code) == 'K0630390':
        return 490
    if f_mode == "関西工場" and line == '4号機':
        return SPEED_4GO.get(str(item_code), SPEED_4GO_DEFAULT)
    if f_mode == "本社": return 400 if line == '2号機' else ((70 if vol == 55 else (100 if vol == 30 else 250)) if line == '3号機' else ((730 if vol in [12, 14] else 650) if line == '5号機' else 260))
    else: return 388 if line == '1号機' else (494 if line == '2号機' else ((70 if vol == 55 else (100 if vol == 30 else 191)) if line == '3号機' else (646 if line == '5号機' else (597 if line == '6号機' else 107))))

# =====================================================================
# 🌟 マスタスタンバイ外側チェック
# =====================================================================

has_local_master = os.path.exists("bom_master_local.csv") or os.path.exists("bom_master.xlsx") or os.path.exists("bom_master.csv") or ('bom_data' in st.session_state) or (file_bom is not None)

has_master_in_gekkan = False
if not has_local_master and file_gekkan is not None:
    try:
        safe_seek(file_gekkan)
        xl_peek = pd.ExcelFile(file_gekkan)
        if any(any(k in s for k in ["マスタ", "BOM", "BomMaster", "ﾏｽﾀ"]) for s in xl_peek.sheet_names):
            has_master_in_gekkan = True
    except: pass

if has_local_master or has_master_in_gekkan:
    st.sidebar.success("🟢 構成表マスタ: 読込済み (スタンバイOK)")
else:
    st.sidebar.warning("⚠️ 構成表マスタが未登録です。")

# 新規ファイルがアップロードされた場合は解析し、既存の学習辞書へ累積マージ
# （上書きではなくマージなので、過去に学習した品目のライン情報は消えない）
if file_jisseki is not None:
    try:
        safe_seek(file_jisseki)
        _df_j_new = pd.read_excel(file_jisseki, header=3)
        _new_dict = parse_jisseki(_df_j_new)
        # 選択中の工場に対応する品目コードのみ受け入れる（本社=H・関西=K）
        # → 工場切替時にアップローダーへ残った他工場のファイルが混入するのを防ぐ
        _pfx = 'H' if factory_mode == "本社" else 'K'
        _new_dict = {k: v for k, v in _new_dict.items() if str(k).startswith(_pfx)}
        if _new_dict:
            _merged = dict(_load_line_dict(factory_mode))
            _merged.update(_new_dict)
            _save_line_dict(_merged, factory_mode)
    except: pass

_cur_line_dict = _load_line_dict(factory_mode)
if _cur_line_dict:
    st.sidebar.success(f"🟢 実績ライン学習（{factory_mode}）: {len(_cur_line_dict)}品目 保持中")
    # 恒久保存用のダウンロードボタン（リポジトリ同梱でアプリ再起動後も維持できる）
    _dict_csv = pd.DataFrame({'品目コード': list(_cur_line_dict.keys()),
                              'ライン': list(_cur_line_dict.values())}).to_csv(index=False).encode('utf-8')
    st.sidebar.download_button(
        "💾 学習辞書CSVをダウンロード",
        _dict_csv,
        file_name=JISSEKI_DICT_FILES.get(factory_mode, "jisseki_line.csv"),
        help="アプリの再起動（コード更新・スリープ）でサーバー上の保存は消えます。このCSVをダウンロードし、GitHubリポジトリのschedule_app.pyと同じ場所に置いてデプロイすると、再起動後も学習内容が恒久的に保持されます。"
    )
else:
    st.sidebar.info("ℹ️ 製造実績レポート未登録（ルールベースで動作）")

# =====================================================================

if st.sidebar.button("🚀 製造計画スケジュールを生成する"):
    if not file_zai or not file_gekkan:
        st.error("エラー: 必要ファイルをアップロードしてください。")
    else:
        with st.spinner("⚡ 裏側でマスタを展開し、エコ・ハッシュエンジンで計画ファイルを出力中..."):
            try:
                df_bom = None
                if file_bom is not None:
                    safe_seek(file_bom)
                    if file_bom.name.endswith('.csv'):
                        try: df_bom = pd.read_csv(file_bom, encoding='utf-8')
                        except:
                            safe_seek(file_bom)
                            df_bom = pd.read_csv(file_bom, encoding='cp932')
                    else: df_bom = clean_bom_master(load_excel_sheets_merged(file_bom, ["マスタ", "BOM", "BomMaster", "ﾏｽﾀ"]))
                elif 'bom_data' in st.session_state:
                    df_bom = st.session_state['bom_data']
                elif os.path.exists("bom_master_local.csv"):
                    try: df_bom = pd.read_csv("bom_master_local.csv", encoding='utf-8')
                    except: df_bom = pd.read_csv("bom_master_local.csv", encoding='cp932')
                elif os.path.exists("bom_master.xlsx"):
                    df_bom = clean_bom_master(pd.read_excel("bom_master.xlsx", header=None))
                elif os.path.exists("bom_master.csv"):
                    try: df_bom = clean_bom_master(pd.read_csv("bom_master.csv", encoding='utf-8', header=None))
                    except: df_bom = clean_bom_master(pd.read_csv("bom_master.csv", encoding='cp932', header=None))

                if df_bom is None and file_gekkan is not None:
                    try:
                        safe_seek(file_gekkan)
                        xl_g = pd.ExcelFile(file_gekkan)
                        m_sheets = [s for s in xl_g.sheet_names if any(k in s for k in ["マスタ", "BOM", "BomMaster", "ﾏｽﾀ"])]
                        if m_sheets:
                            df_bom = clean_bom_master(pd.read_excel(xl_g, sheet_name=m_sheets[0], header=None))
                    except: pass

                if df_bom is not None:
                    try:
                        df_bom.to_csv("bom_master_local.csv", index=False, encoding='utf-8')
                        st.session_state['bom_data'] = df_bom
                    except: pass

                if df_bom is None:
                    st.error("エラー: 構成表マスタが見つかりません。")
                    st.stop()

                bom_lookup_dict = {}
                if not df_bom.empty:
                    p_col = next((c for c in df_bom.columns if c in [
                        '親品目コード', '商品CODE', '商品コード', '品目コード', '商品CD'
                    ]), df_bom.columns[0])
                    c_col = next((c for c in df_bom.columns if c in [
                        '子品目コード', '配合CODE', '配合コード', '配合CD', '中身コード'
                    ]), df_bom.columns[1])

                    for _, r in df_bom.iterrows():
                        pv_raw = str(r[p_col]).strip()
                        cv = str(r[c_col]).strip()
                        if '.' in pv_raw: pv_raw = pv_raw.split('.')[0]
                        pv_digits = "".join(re.findall(r'\d+', pv_raw))
                        pv_prefix_match = re.match(r'^([A-Za-z]+)', pv_raw)
                        pv_prefix = pv_prefix_match.group(1).upper() if pv_prefix_match else ''
                        pv_clean = f"{pv_prefix}_{pv_digits}" if pv_digits else ''
                        if pv_clean:
                            existing = bom_lookup_dict.get(pv_clean)
                            # 優先順位1: BK・BHプレフィックスは最優先で上書き
                            if cv.startswith(('BK', 'BH')):
                                bom_lookup_dict[pv_clean] = cv
                            # 優先順位2: K・H始まり末尾1は無視
                            elif (cv.startswith('K') or cv.startswith('H')) and cv.endswith('1'):
                                continue
                            # 優先順位3: K・H始まり末尾0はBK/BHがない場合に選択
                            elif (cv.startswith('K') or cv.startswith('H')) and cv.endswith('0'):
                                if existing is None or not existing.startswith(('BK', 'BH')):
                                    bom_lookup_dict[pv_clean] = cv
                            # それ以外: まだ何も登録がない場合のみ
                            else:
                                if existing is None:
                                    bom_lookup_dict[pv_clean] = cv

                # --- 品目マスタ（袋サイズ切り替えルール用）読み込み ---
                # 優先順位: ①手動アップロード(明示的な上書き) → ②MRPアプリと共有のGitHubリポジトリから自動取得
                #           → ③セッションキャッシュ → ④ローカルキャッシュ
                df_itemmaster = None
                itemmaster_source = None
                if file_itemmaster is not None:
                    safe_seek(file_itemmaster)
                    if file_itemmaster.name.endswith('.csv'):
                        try: df_itemmaster = pd.read_csv(file_itemmaster, encoding='utf-8')
                        except:
                            safe_seek(file_itemmaster)
                            df_itemmaster = pd.read_csv(file_itemmaster, encoding='cp932')
                    else:
                        df_itemmaster = pd.read_excel(file_itemmaster)
                    itemmaster_source = "手動アップロード"
                else:
                    _im_bytes = fetch_github_file_bytes("data/ItemMaster.csv")
                    if _im_bytes:
                        try:
                            df_itemmaster = read_item_master_bytes(_im_bytes, "ItemMaster.csv")
                            itemmaster_source = "MRPアプリ共有マスタ（GitHub自動取得）"
                        except Exception:
                            df_itemmaster = None
                    if df_itemmaster is None and 'itemmaster_data' in st.session_state:
                        df_itemmaster = st.session_state['itemmaster_data']
                        itemmaster_source = "セッションキャッシュ"
                    if df_itemmaster is None and os.path.exists("itemmaster_local.csv"):
                        try: df_itemmaster = pd.read_csv("itemmaster_local.csv", encoding='utf-8')
                        except: df_itemmaster = pd.read_csv("itemmaster_local.csv", encoding='cp932')
                        itemmaster_source = "ローカルキャッシュ"

                if df_itemmaster is not None:
                    try:
                        df_itemmaster.to_csv("itemmaster_local.csv", index=False, encoding='utf-8')
                        st.session_state['itemmaster_data'] = df_itemmaster
                    except: pass
                    st.caption(f"📦 品目マスタ（袋サイズ判定用）: {itemmaster_source} — {len(df_itemmaster)}件")
                else:
                    st.caption("ℹ️ 品目マスタ未取得のため、袋サイズ切り替えルールは無効です（容量ベースの従来ロジックで動作）。")

                item_bagsize_dict = {}
                if df_itemmaster is not None and not df_itemmaster.empty:
                    im_code_col = next((c for c in df_itemmaster.columns if c in ['品目コード', '商品コード', '商品CODE']), None)
                    im_size_col = next((c for c in df_itemmaster.columns if '幅' in str(c) and 'ピッチ' in str(c)), None)
                    if im_code_col and im_size_col:
                        for _, r in df_itemmaster.iterrows():
                            code = str(r[im_code_col]).strip()
                            bag = parse_bag_size(r[im_size_col])
                            if code and bag:
                                item_bagsize_dict[code] = bag

                def extract_content_code(item_code):
                    item_str = str(item_code).strip()
                    if '.' in item_str: item_str = item_str.split('.')[0]
                    item_digits = "".join(re.findall(r'\d+', item_str))
                    item_prefix_match = re.match(r'^([A-Za-z]+)', item_str)
                    item_prefix = item_prefix_match.group(1).upper() if item_prefix_match else ''
                    item_clean = f"{item_prefix}_{item_digits}" if item_digits else ''
                    return bom_lookup_dict.get(item_clean, item_code)

                # --- 在庫推移リスト読み込み ---
                df_zai_raw = load_excel_sheets_merged(file_zai, ["在庫推移リスト", "在庫推移"])
                header_idx = next((i for i in range(len(df_zai_raw)) if any(kw in [str(v).strip() for v in df_zai_raw.iloc[i].values] for kw in ['品目コード', '品目ｺｰﾄﾞ', '商品コード', '商品CD', '商品CODE'])), None)
                if header_idx is None: st.error("エラー: 見出し列が見つかりません。"); st.stop()

                raw_headers = [str(h).strip() for h in df_zai_raw.iloc[header_idx].values]
                standard_headers = ['品目コード' if h in ['品目コード', '品目ｺｰﾄﾞ', '商品コード', '商品CD', '商品CODE'] else ('品目名' if h in ['品目名', '商品名'] else ('安全在庫数' if h in ['安全在庫数', '安全在庫'] else ('種類' if h in ['種類', '区分'] else h))) for h in raw_headers]

                df_zai_fixed = df_zai_raw.iloc[header_idx+1:].copy()
                df_zai_fixed.columns = standard_headers
                df_zai_fixed['品目コード'] = df_zai_fixed['品目コード'].ffill().astype(str).str.strip()
                df_zai_fixed['品目名'] = df_zai_fixed['品目名'].ffill().astype(str).str.strip()
                df_zai_fixed['安全在庫数'] = df_zai_fixed['安全在庫数'].ffill()

                df_zai_in_zai = df_zai_fixed[df_zai_fixed['種類'] == '在'].copy()
                df_zai_in_zai['安全在庫数'] = pd.to_numeric(df_zai_in_zai['安全在庫数'], errors='coerce')
                date_col_pattern = re.compile(r'^\d{1,2}[/\-]\d{1,2}\(.\)$')
                date_columns = [c for c in df_zai_in_zai.columns if date_col_pattern.match(str(c).strip())]
                base_date = date_columns[0] if date_columns else df_zai_in_zai.columns[-1]
                df_zai_in_zai['現在の在庫'] = pd.to_numeric(df_zai_in_zai[base_date], errors='coerce')

                # --- 確定済み入庫予定（入）の合計を計算 ---
                iko_total_dict = {}
                iko_entries = []
                if consider_iko:
                    df_zai_in_iri = df_zai_fixed[df_zai_fixed['種類'] == '入'].copy()
                    if not df_zai_in_iri.empty and date_columns:
                        for c in date_columns:
                            df_zai_in_iri[c] = pd.to_numeric(df_zai_in_iri[c], errors='coerce').fillna(0)
                        df_zai_in_iri['入庫予定合計'] = df_zai_in_iri[date_columns].sum(axis=1)
                        iko_total_dict = df_zai_in_iri.groupby('品目コード')['入庫予定合計'].sum().to_dict()

                        def _parse_date_col(col_name, base_today):
                            m_dt = re.match(r'^(\d{1,2})[/\-](\d{1,2})\(', str(col_name).strip())
                            if not m_dt:
                                return None
                            mo, da = int(m_dt.group(1)), int(m_dt.group(2))
                            y = base_today.year
                            if mo < base_today.month - 6:
                                y += 1
                            try:
                                return datetime.date(y, mo, da)
                            except: return None

                        _today_for_parse = datetime.date.today()
                        for _, irow in df_zai_in_iri.iterrows():
                            i_code = str(irow['品目コード']).strip()
                            for c in date_columns:
                                qty = irow[c]
                                if qty and qty > 0:
                                    d_parsed = _parse_date_col(c, _today_for_parse)
                                    if d_parsed:
                                        iko_entries.append({'品目コード': i_code, '日付': d_parsed, '数量': int(qty)})

                df_zai_in_zai['入庫予定合計'] = df_zai_in_zai['品目コード'].map(iko_total_dict).fillna(0.0)
                df_zai_in_zai['安全割れ不足数'] = (df_zai_in_zai['安全在庫数'] - df_zai_in_zai['現在の在庫'] - df_zai_in_zai['入庫予定合計']).apply(lambda x: max(0, x))

                # --- 月間計画書読み込み ---
                _gekkan_keywords = ["本社 月間製造計画書", "月間製造計画書", "月間計画", "本社"] if factory_mode == "本社" else ["関西工場 月間製造計画書", "関西工場", "関西製造計画", "計画", "月間製造計画書"]
                df_monthly_raw = load_excel_sheets_merged(
                    file_gekkan,
                    _gekkan_keywords,
                    exclude_keywords=["天川"] if factory_mode == "本社" else ["外製", "参考"]
                )
                # 読み込んだシートを画面に表示（複数シートが正しく読まれているかの確認用）
                _sheets_read = getattr(load_excel_sheets_merged, 'last_matched', [])
                st.info(f"📄 月間製造計画書の読込シート（{len(_sheets_read)}枚）: {'、'.join(_sheets_read) if _sheets_read else '不明'}")
                if factory_mode == "関西工場" and len(_sheets_read) < 2:
                    st.warning("⚠️ 関西工場の月間製造計画書は2シートの読み込みが想定されています。1シートのみの場合は、シート名に「関西」「計画」等のキーワードが含まれているかご確認ください。")
                item_row_idx = next((i for i in range(min(15, len(df_monthly_raw))) if any(kw in [str(v).strip() for v in df_monthly_raw.iloc[i].values] for kw in ['商品CD', '商品コード', '品目コード', '品目ｺｰﾄﾞ', '品目ｃｄ', '商品CODE'])), 1)

                target_month_num = None
                try:
                    target_month_num = int(target_month.replace("月", ""))
                except: pass

                # 月ラベル列の位置を全てスキャン
                # （タイトル行〜品目コード見出し行「自体」まで全行・部分一致で頑健に検出。
                #   見出し行に「7月度予定」等の形で月が埋め込まれるレイアウトにも対応）
                month_label_cols = []
                _seen_cols = set()
                for r_scan in range(0, item_row_idx + 1):
                    for c_scan in range(df_monthly_raw.shape[1]):
                        if c_scan in _seen_cols:
                            continue
                        cell_val = df_monthly_raw.iloc[r_scan, c_scan]
                        if cell_val is not None:
                            cell_str = str(cell_val).strip()
                            # 「7月」「7月度」「2026年7月」「7月度予定」等を許容
                            m = re.search(r'(\d{1,2})\s*月', cell_str)
                            if m and not re.search(r'\d{1,2}\s*月\s*\d{1,2}\s*日', cell_str):
                                mn = int(m.group(1))
                                if 1 <= mn <= 12:
                                    month_label_cols.append((c_scan, mn))
                                    _seen_cols.add(c_scan)
                month_label_cols.sort(key=lambda x: x[0])

                # 連続する同一月ラベルを1つにまとめる
                # （「7月度予定」「7月度実績」が並ぶ見出し行埋め込み型で、
                #   月ブロックが1列幅に縮まないようにするため）
                _dedup = []
                for col_pos, m_num in month_label_cols:
                    if _dedup and _dedup[-1][1] == m_num:
                        continue  # 直前と同じ月ならスキップ（ブロック開始位置は最初の列）
                    _dedup.append((col_pos, m_num))
                month_label_cols = _dedup

                def _find_month_block(month_num):
                    """指定月の(ブロック開始列, 終了列)を返す"""
                    for idx_lbl, (col_pos, m_num) in enumerate(month_label_cols):
                        if m_num == month_num:
                            b_end = month_label_cols[idx_lbl + 1][0] if idx_lbl + 1 < len(month_label_cols) else df_monthly_raw.shape[1]
                            return col_pos, b_end
                    return None, None

                # 予定・実績の見出しは「商品CODE行」自体か「その1つ下の行」にあるレイアウトの両方に対応
                _hdr_scan_rows = [item_row_idx]
                if item_row_idx + 1 < len(df_monthly_raw):
                    _hdr_scan_rows.append(item_row_idx + 1)
                _hdr_rows_vals = {r: [str(v).strip() for v in df_monthly_raw.iloc[r].values] for r in _hdr_scan_rows}

                def _find_plan_actual_cols(b_start, b_end):
                    """ブロック内の予定・実績列インデックスを返す。
                    「製造予定」「製造実績」の完全一致を最優先し、次に部分一致、
                    どちらも無い場合のみブロック位置から推定する。"""
                    rng = list(range(b_start, b_end)) if b_start is not None else list(range(df_monthly_raw.shape[1]))
                    # パス1: 「製造予定」「製造実績」の厳密一致（月末在庫・出荷予測との誤認防止）
                    for _hr in _hdr_scan_rows:
                        vals = _hdr_rows_vals[_hr]
                        p_idx = next((c for c in rng if c < len(vals) and '製造予定' in vals[c]), None)
                        a_idx = next((c for c in rng if c < len(vals) and '製造実績' in vals[c]), None)
                        if p_idx is not None and a_idx is not None:
                            return p_idx, a_idx
                    # パス2: 「予定」「計画」/「実績」の部分一致（「予測」は除外）
                    for _hr in _hdr_scan_rows:
                        vals = _hdr_rows_vals[_hr]
                        p_idx = None; a_idx = None
                        for c in rng:
                            t = vals[c] if c < len(vals) else ''
                            if ('予定' in t or '計画' in t) and '予測' not in t and p_idx is None: p_idx = c
                            elif '実績' in t and a_idx is None: a_idx = c
                        if p_idx is not None and a_idx is not None:
                            return p_idx, a_idx
                        if p_idx is not None and a_idx is None and b_start is not None and p_idx + 1 < b_end:
                            return p_idx, p_idx + 1  # 予定の右隣を実績とみなす
                    # パス3: ブロック先頭2列を予定・実績とみなす（最終手段）
                    if b_start is not None and b_end - b_start >= 2:
                        return b_start, b_start + 1
                    return None, None

                # 対象月リストを決定（期末一括なら対象月〜10月度、通常は対象月のみ）
                if plan_to_yearend and target_month_num is not None:
                    if target_month_num <= 10:
                        months_to_plan = list(range(target_month_num, 11))
                    else:
                        # 11月・12月始まりは翌期の10月度まで
                        months_to_plan = list(range(target_month_num, 13)) + list(range(1, 11))
                else:
                    months_to_plan = [target_month_num] if target_month_num is not None else []

                _first_month = months_to_plan[0] if months_to_plan else target_month_num
                _month_order = {m: i for i, m in enumerate(months_to_plan)}

                # 各月の予定・実績列を収集
                month_col_pairs = []
                for m_num in months_to_plan:
                    b_start, b_end = _find_month_block(m_num)
                    p_idx, a_idx = _find_plan_actual_cols(b_start, b_end)
                    if p_idx is not None and a_idx is not None:
                        month_col_pairs.append((m_num, p_idx, a_idx))

                # フォールバック（1つも見つからない場合）
                if not month_col_pairs:
                    try:
                        p_fb = 46 + (int(target_month.replace("月", "")) - 6) * 2
                        month_col_pairs = [(target_month_num, p_fb, p_fb + 1)]
                    except:
                        month_col_pairs = [(target_month_num, 46, 47)]

                code_col_idx = next((c for c in range(min(5, len(df_monthly_raw.columns))) if any(kw in str(df_monthly_raw.iloc[item_row_idx, c]) for kw in ['商品CD', '品目コード', 'コード', '商品', '商品CODE'])), 0)
                name_col_idx = next((c for c in range(min(5, len(df_monthly_raw.columns))) if any(kw in str(df_monthly_raw.iloc[item_row_idx, c]) for kw in ['商品名', '品目名', '名', '品名'])), 1)

                # 生産ロット列（CA列基準）の検出：見出しに「ロット」を含む列を優先、無ければCA列(79列目)
                _lot_col_idx = None
                for _hr in _hdr_scan_rows:
                    _vals_l = _hdr_rows_vals[_hr]
                    _lot_col_idx = next((c for c, t in enumerate(_vals_l) if 'ロット' in t or 'ﾛｯﾄ' in t), None)
                    if _lot_col_idx is not None: break
                if _lot_col_idx is None and df_monthly_raw.shape[1] > 78:
                    _lot_col_idx = 78  # CA列
                lot_dict = {}
                if _lot_col_idx is not None:
                    for _li in range(item_row_idx + 1, len(df_monthly_raw)):
                        _lc = str(df_monthly_raw.iloc[_li, code_col_idx]).strip()
                        _lv = pd.to_numeric(df_monthly_raw.iloc[_li, _lot_col_idx], errors='coerce')
                        if _lc and _lc not in ('nan', 'None', '') and not pd.isna(_lv) and _lv > 0 and _lc not in lot_dict:
                            lot_dict[_lc] = float(_lv)
                if lot_dict:
                    st.info(f"📦 生産ロット指定（CA列）を検出: {len(lot_dict)}品目。ロット記載品はロット単位で切上げ、未記載品は従来のm3バッチで計算します。")

                df_m = df_monthly_raw.iloc[item_row_idx+1:].copy()
                # 月ごとの計画残数（予定−実績、マイナスは0）を月別に保持しつつ合計も計算する
                _total_zan = None
                _total_plan = None
                _total_actual = None
                _month_zan = {}
                for (m_num, p_idx, a_idx) in month_col_pairs:
                    _p = pd.to_numeric(df_m.iloc[:, p_idx], errors='coerce').fillna(0)
                    _a = pd.to_numeric(df_m.iloc[:, a_idx], errors='coerce').fillna(0)
                    _z = (_p - _a).clip(lower=0)
                    _month_zan[m_num] = _z
                    _total_zan    = _z if _total_zan is None else _total_zan + _z
                    _total_plan   = _p if _total_plan is None else _total_plan + _p
                    _total_actual = _a if _total_actual is None else _total_actual + _a

                df_m_clean = pd.DataFrame({
                    '品目コード': df_m.iloc[:, code_col_idx].astype(str).str.strip(),
                    '品目名_計画書': df_m.iloc[:, name_col_idx].astype(str).str.strip(),
                    '選択月_製造予定': _total_plan.values,
                    '選択月_製造実績': _total_actual.values
                })
                df_m_clean['選択月_計画残数'] = _total_zan.values
                for _mz_num, _mz_series in _month_zan.items():
                    df_m_clean[f'残数_{_mz_num}'] = _mz_series.values
                if plan_to_yearend:
                    _found_months = [m for m, _, _ in month_col_pairs]
                    _missing_months = [m for m in months_to_plan if m not in _found_months]
                    _m_names = "、".join(f"{m}月度" for m in _found_months)
                    st.info(f"📆 期末一括計画モード: {_m_names} を各月度ごとに区切って計画します（各月の計画はその月内でスケジュールされます）。")
                    if _missing_months:
                        _miss_names = "、".join(f"{m}月度" for m in _missing_months)
                        # 診断用: スキャンした行の内容サンプルを収集
                        _diag_rows = []
                        for _r_diag in range(0, min(item_row_idx + 1, 5)):
                            _row_sample = [str(v).strip() for v in df_monthly_raw.iloc[_r_diag].values[:30] if str(v).strip() not in ('nan', 'None', '')]
                            _diag_rows.append(f"行{_r_diag + 1}: {_row_sample[:15]}")
                        st.warning(
                            f"⚠️ {_miss_names} の計画列が月間計画書から検出できませんでした。"
                            f"（検出済み月ラベル位置: {[(c, f'{m}月') for c, m in month_label_cols]} / 見出し行: {item_row_idx + 1}行目）"
                        )
                        with st.expander("🔍 診断情報（スキャンした行の内容）"):
                            for _dr in _diag_rows:
                                st.text(_dr)
                df_m_distinct = df_m_clean[df_m_clean['品目コード'].notna() & (~df_m_clean['品目コード'].isin(['nan', '', 'None']))].drop_duplicates(subset=['品目コード'])

                del df_monthly_raw, df_m, df_m_clean
                gc.collect()

                # =====================================================================
                # 🌟 1期間（1か月度）分の計画を独立して立案する関数
                #    use_inventory: 在庫・安全在庫の考慮（初月のみTrue）
                #    use_confirmed: 確定入庫予定の反映（初月のみTrue）
                # =====================================================================
                def _plan_period(df_m_period, use_inventory, use_confirmed, period_start, target_bd, day_offset, month_label):
                    all_codes = (set(df_zai_in_zai['品目コード']) if use_inventory else set()).union(set(df_m_period[df_m_period['選択月_計画残数'] > 0]['品目コード']))
                    master_list = []
                    for code in all_codes:
                        EXCLUDE_CODES = {
                            '合計', 'nan', '商品CD', '品目コード', 'None', '商品CODE',
                            '進捗状況', '製造進捗状況', '小計', '総合計', '合　計',
                            '品目ｺｰﾄﾞ', '品目cd', '品目ｃｄ', ''
                        }
                        if code in EXCLUDE_CODES or not re.search(r'\d', str(code)) or (factory_mode == "関西工場" and str(code).startswith('H')): continue
                        zai_row = df_zai_in_zai[df_zai_in_zai['品目コード'] == code] if use_inventory else df_zai_in_zai.iloc[0:0]
                        plan_row = df_m_period[df_m_period['品目コード'] == code]
                        _common = {
                            '品目コード': code,
                            '品目名': zai_row['品目名'].iloc[0] if not zai_row.empty else (plan_row['品目名_計画書'].iloc[0] if not plan_row.empty else "不明"),
                            '現在の在庫': zai_row['現在の在庫'].iloc[0] if not zai_row.empty else np.nan,
                            '安全在庫数': zai_row['安全在庫数'].iloc[0] if not zai_row.empty else np.nan,
                        }
                        _anzen = zai_row['安全割れ不足数'].iloc[0] if not zai_row.empty else 0.0

                        def _zan_of(m_q):
                            _cn = f'残数_{m_q}'
                            if not plan_row.empty and _cn in plan_row.columns:
                                return float(plan_row[_cn].iloc[0])
                            if m_q == _first_month and not plan_row.empty:
                                return float(plan_row['選択月_計画残数'].iloc[0])
                            return 0.0

                        # 月度ごとに独立した計画行を作成（先頭月は安全在庫要素も含む）
                        for m_num in months_to_plan:
                            zan_m = _zan_of(m_num)
                            if m_num == _first_month:
                                master_list.append({**_common, '安全割れ不足数': _anzen, '今月の計画残数': zan_m, '対象月度': m_num})
                            elif zan_m > 0:
                                master_list.append({**_common, '安全割れ不足数': 0.0, '今月の計画残数': zan_m, '対象月度': m_num})

                    df_master_combined = pd.DataFrame(master_list)
                    if df_master_combined.empty: return [], pd.DataFrame(), 0, 0

                    df_master_combined['採用ベース数量'] = df_master_combined[['安全割れ不足数', '今月の計画残数']].max(axis=1)
                    df_master_combined = df_master_combined[df_master_combined['採用ベース数量'] > 0].copy()

                    df_master_combined['容量_L'] = df_master_combined['品目名'].apply(extract_volume_safe)
                    df_master_combined['kg品フラグ'] = df_master_combined['品目名'].apply(is_kg_product)
                    df_master_combined['kg重量'] = df_master_combined['品目名'].apply(get_kg_weight)
                    df_master_combined['ベース必要容量_L'] = df_master_combined.apply(
                        lambda r: r['採用ベース数量'] * r['kg重量'] if r['kg品フラグ'] else r['採用ベース数量'] * max(r['容量_L'], 1),
                        axis=1
                    )
                    df_master_combined['中身設計コード'] = df_master_combined['品目コード'].apply(extract_content_code)
                    df_master_combined['生産ロット'] = df_master_combined['品目コード'].map(lot_dict).fillna(0.0)

                    def calc_batch(group_df):
                        name = group_df['品目名'].iloc[0]
                        is_kasei = '化成肥料' in name and 'ｺｰﾅﾝ' in name
                        # kg品のみのグループは重量ベース（1000kg単位）でバッチを計算する
                        # （m3単位と混同すると袋数がほぼ0になり計画から脱落するため）
                        is_all_kg = bool(group_df['kg品フラグ'].all()) if 'kg品フラグ' in group_df.columns else False
                        total = group_df['ベース必要容量_L'].sum()
                        if is_kasei or is_all_kg:
                            batches = math.ceil(total / 1000)
                            return float(batches * 1000)
                        else:
                            m3 = (total / 0.9) / 1000
                            return 5.0 if m3 <= 5.0 else (10.0 if m3 <= 10.0 else float(math.ceil(m3 / 10.0) * 10.0))

                    # ロット記載品は個別にロット単位で切上げるため、m3バッチ集計から除外する
                    _df_batch_src = df_master_combined[df_master_combined['生産ロット'] <= 0]
                    if not _df_batch_src.empty:
                        grouped = _df_batch_src.groupby(['中身設計コード', '対象月度']).apply(
                            lambda g: pd.Series({'ベース必要容量_L': g['ベース必要容量_L'].sum(),
                                                 '製造決定_m3': calc_batch(g),
                                                 'kg品フラグ': g['kg品フラグ'].any()})
                        ).reset_index()
                    else:
                        grouped = pd.DataFrame(columns=['中身設計コード', '対象月度', 'ベース必要容量_L', '製造決定_m3', 'kg品フラグ'])

                    df_final = df_master_combined.merge(grouped[['中身設計コード', '対象月度', '製造決定_m3', 'kg品フラグ']], on=['中身設計コード', '対象月度'], how='left', suffixes=('', '_g'))
                    df_final['製造決定_m3'] = df_final['製造決定_m3'].fillna(0.0)

                    total_vol_recipe = df_final.groupby(['中身設計コード', '対象月度'])['ベース必要容量_L'].transform('sum')
                    df_final['分配比率'] = (df_final['ベース必要容量_L'] / total_vol_recipe).fillna(1.0)

                    KEYWORDS_4GO_LOT = ['ピートモス', 'くん炭', 'バーミキュライト', 'パーライト',
                                        'ﾋﾟｰﾄﾓｽ', 'ﾊﾞｰﾐｷｭﾗｲﾄ', 'ﾊﾟｰﾗｲﾄ']

                    def calc_bags(r):
                        # 生産ロット記載品: 計画数量をロット単位で切上げ（CA列基準）
                        _lot = r.get('生産ロット', 0)
                        if _lot and _lot > 0:
                            _base_q = r['採用ベース数量']
                            return int(math.ceil(_base_q / _lot) * _lot) if _base_q > 0 else 0
                        name_str = str(r['品目名'])
                        if any(k in name_str for k in KEYWORDS_4GO_LOT):
                            base_qty = r['採用ベース数量']
                            return int(math.ceil(base_qty / 10.0) * 10) if base_qty > 0 else 0
                        if 'CLEAR' in name_str and 'ERA' in name_str:
                            return int(r['採用ベース数量'])
                        if r['kg品フラグ_g'] if 'kg品フラグ_g' in r else r.get('kg品フラグ', False):
                            unit_kg = r['kg重量'] if r['kg重量'] > 0 else 1.0
                            return int(round((r['製造決定_m3'] * r['分配比率']) / unit_kg))
                        else:
                            vol = max(r['容量_L'], 1)
                            return int(round((r['製造決定_m3'] * 1000 * 0.9 * r['分配比率']) / vol))

                    df_final['計画製造袋数'] = df_final.apply(calc_bags, axis=1).clip(lower=0)
                    df_final['製造理由'] = df_final.apply(lambda r: ('計画未達' if r.get('対象月度', _first_month) != _first_month else ('現在庫がマイナス' if not pd.isna(r['現在の在庫']) and r['現在の在庫'] < 0 else ('安全在庫割れ' if r['安全割れ不足数'] > 0 else '計画未達'))), axis=1)
                    df_final['計画製造袋数'] = df_final.apply(lambda r: 0 if r['製造理由'] == '計画未達' and r['計画製造袋数'] < 100 and r.get('生産ロット', 0) <= 0 else r['計画製造袋数'], axis=1)
                    df_final['堆肥・腐葉土フラグ'] = df_final['品目名'].apply(lambda n: any(k in str(n) for k in ['腐葉土', '堆肥', '特大袋']))

                    # =====================================================================
                    # 🌟 ライン割り当て
                    # =====================================================================
                    if factory_mode == "本社":
                        # 本社ライン割り当てルール：
                        #   固定コード            → 6号機
                        #   堆肥・腐葉土・再生材系 → 3号機
                        #   9L以下               → その他（外注）
                        #   10L〜25L             → 5号機（速度優先）
                        #   26L                  → 2号機（速度優先）
                        #   27L以上              → 6号機
                        df_final['製造ライン'] = df_final.apply(lambda r:
                            '6号機' if r['品目コード'] in ('H0690020', 'H0690000', 'H0690030', 'H0390000')
                            else (HONSHA_FIXED_LINE[r['品目コード']] if r['品目コード'] in HONSHA_FIXED_LINE
                            else ('3号機' if r['品目コード'] == 'H0620030' or any(k in r['品目名'] for k in ['再生材', 'もう一土元気']) or r['堆肥・腐葉土フラグ']
                            else ('その他' if r['容量_L'] < 10
                            else ('5号機' if r['容量_L'] <= 25
                            else ('2号機' if r['容量_L'] <= 26
                            else '6号機'))))), axis=1)
                        # 同一配合サイズ違いを可能な限り同一ラインに統一
                        df_final = unify_recipe_lines(df_final, factory_mode)
                    else:
                        recipe_total_bags = df_master_combined.groupby('中身設計コード')['採用ベース数量'].sum().to_dict() if '中身設計コード' in df_master_combined.columns else {}

                        # 学習辞書（累積マージ済み）を使用
                        jisseki_line_dict = dict(_load_line_dict(factory_mode))

                        KEYWORDS_4GO = ['ピートモス', 'くん炭', 'バーミキュライト', 'パーライト',
                                        'ﾋﾟｰﾄﾓｽ', 'ﾊﾞｰﾐｷｭﾗｲﾄ', 'ﾊﾟｰﾗｲﾄ']

                        def assign_line_kansai(r):
                            name = r['品目名']
                            vol = r['容量_L']
                            code = r['品目コード']
                            is_compost = r['堆肥・腐葉土フラグ']
                            is_special = any(k in name for k in ['再生材', 'もう一土元気'])

                            if any(k in name for k in KEYWORDS_4GO):
                                return '4号機'

                            FIXED_CODES_SONOTA = ('K0270450', 'K0190010')

                            if code in ('K0390110', 'K0480080', 'K0680190'):
                                return '3号機'
                            if code in FIXED_CODES_SONOTA:
                                return 'その他'
                            if code == 'K0430120':
                                return '4号機'
                            if 'CLEAR' in name and 'ERA' in name:
                                return 'その他'
                            if '有機石灰' in name:
                                return 'その他'
                            if is_kg_product(name) and get_kg_weight(name) < 1.0:
                                return '4号機'

                            is_compost_only = any(k in name for k in ['腐葉土', '堆肥'])
                            if is_compost or is_special:
                                if '特大袋' in name:
                                    return '3号機'
                                elif vol >= 40:
                                    return '1号機'
                                elif vol >= 14:
                                    return '3号機'
                                elif is_compost_only:
                                    return '4号機'
                                elif vol >= 1.2:
                                    return '5号機'
                                else:
                                    return 'その他'

                            if '化成肥料' in name and 'ｺｰﾅﾝ' in name:
                                return '5号機'
                            if code == 'K0630390':
                                return '5号機'

                            if code in jisseki_line_dict:
                                return jisseki_line_dict[code]

                            if str(code).startswith('K0225') and '専用培養土' in name and '12' in name:
                                return '5号機'

                            # kg品（vol < 0）の容量換算による割り当て
                            # 関西工場：10〜20L → 6号機、21〜25L → 2号機、26L以上 → 1号機（速度優先）
                            if vol < 0:
                                kg_w = get_kg_weight(name)
                                if kg_w < 1.0:
                                    return '4号機'
                                else:
                                    eff_vol = int(kg_w)
                                    if eff_vol < 10:
                                        return '5号機'
                                    elif eff_vol >= 26:
                                        return '1号機'
                                    elif eff_vol <= 20:
                                        return '6号機'  # 10〜20L：6号機が速い(597 > 494)
                                    else:
                                        return '2号機'  # 21〜25L：2号機

                            # 通常品の容量による割り当て
                            # 関西工場：9L以下 → 5号機、10〜20L → 6号機、21〜25L → 2号機、26L以上 → 1号機（速度優先）
                            if vol < 1.2:
                                return '4号機'
                            if vol < 10:
                                return '5号機'
                            if vol >= 26:
                                return '1号機'
                            if vol <= 20:
                                return '6号機'  # 10〜20L：6号機が速い(597 > 494)
                            return '2号機'      # 21〜25L：2号機

                        df_final['製造ライン'] = df_final.apply(assign_line_kansai, axis=1)
                        # 同一配合サイズ違いを可能な限り同一ラインに統一（物理的に不可能な場合は維持）
                        df_final = unify_recipe_lines(df_final, factory_mode)
                        # 固定コードを統一処理後に強制上書き（unify_recipe_linesによる上書きを防ぐ）
                        FIXED_LINE_KANSAI = {
                            'K0390110': '3号機', 'K0480080': '3号機', 'K0680190': '3号機',
                            'K0270450': 'その他', 'K0190010': 'その他',
                            'K0430120': '4号機', 'K0630390': '5号機',
                        }
                        for fix_code, fix_line in FIXED_LINE_KANSAI.items():
                            df_final.loc[df_final['品目コード'] == fix_code, '製造ライン'] = fix_line

                    # =====================================================================
                    # 🌟 同一配合・ライン違いの場合はライン毎に最小バッチを独立再計算
                    # =====================================================================
                    def recalc_batch_per_line(df):
                        """同一配合でライン違いの品目群について、
                        ライン毎に独立した最小バッチ・分配比率・袋数を再計算する"""
                        df = df.copy()
                        recipe_line_counts = df.groupby(['中身設計コード', '対象月度'])['製造ライン'].nunique()
                        multi_line_recipes = recipe_line_counts[recipe_line_counts > 1].index.tolist()

                        for recipe_code in multi_line_recipes:
                            recipe_mask = (df['中身設計コード'] == recipe_code[0]) & (df['対象月度'] == recipe_code[1])
                            for line, line_grp in df[recipe_mask].groupby('製造ライン'):
                                # ロット記載品はライン別再バッチの対象外
                                if '生産ロット' in line_grp.columns:
                                    line_grp = line_grp[line_grp['生産ロット'] <= 0]
                                if line_grp.empty:
                                    continue
                                name_str = str(line_grp['品目名'].iloc[0])
                                is_kasei = '化成肥料' in name_str and 'ｺｰﾅﾝ' in name_str
                                is_all_kg = bool(line_grp['kg品フラグ'].all()) if 'kg品フラグ' in line_grp.columns else False
                                total_vol = line_grp['ベース必要容量_L'].sum()

                                if is_kasei or is_all_kg:
                                    new_m3 = float(math.ceil(total_vol / 1000) * 1000)
                                else:
                                    m3 = (total_vol / 0.9) / 1000
                                    new_m3 = 5.0 if m3 <= 5.0 else (10.0 if m3 <= 10.0 else float(math.ceil(m3 / 10.0) * 10.0))

                                for idx in line_grp.index:
                                    row_vol = df.loc[idx, 'ベース必要容量_L']
                                    new_ratio = row_vol / total_vol if total_vol > 0 else 1.0
                                    df.loc[idx, '製造決定_m3'] = new_m3
                                    df.loc[idx, '分配比率'] = new_ratio
                        return df

                    df_final = recalc_batch_per_line(df_final)
                    # バッチ再計算後に袋数を更新
                    df_final['計画製造袋数'] = df_final.apply(calc_bags, axis=1).clip(lower=0)
                    df_final['計画製造袋数'] = df_final.apply(
                        lambda r: 0 if r['製造理由'] == '計画未達' and r['計画製造袋数'] < 100 and r.get('生産ロット', 0) <= 0 else r['計画製造袋数'], axis=1
                    )
                    # ロット品の決定m3表示を実数換算（表示用）
                    def _m3_disp(r):
                        if r.get('生産ロット', 0) > 0 and r['計画製造袋数'] > 0:
                            if r.get('kg品フラグ', False):
                                return round(r['計画製造袋数'] * r['kg重量'], 1)
                            return round(r['計画製造袋数'] * max(r['容量_L'], 1) / 900.0, 1)
                        return r['製造決定_m3']
                    df_final['製造決定_m3'] = df_final.apply(_m3_disp, axis=1)

                    df_final['製造所要時間_分'] = df_final.apply(lambda r: (r['計画製造袋数'] / get_sp(r['製造ライン'], r['容量_L'], factory_mode, r['品目コード'])) * 60 if r['計画製造袋数'] > 0 else 0.0, axis=1)
                    df_final['緊急度'] = df_final.apply(lambda r: (r['現在の在庫'] - r['安全在庫数']) if not pd.isna(r['現在の在庫']) else 500, axis=1)
                    df_final['グループ緊急度'] = df_final['中身設計コード'].map(df_final.groupby('中身設計コード')['緊急度'].min().to_dict())
                    df_final['袋サイズ'] = df_final['品目コード'].astype(str).str.strip().apply(lambda c: item_bagsize_dict.get(c))

                    df_final['_月順'] = df_final['対象月度'].map(lambda m: _month_order.get(m, 99))
                    df_final_sorted = df_final[df_final['計画製造袋数'] > 0].sort_values(by=['製造ライン', '_月順', 'グループ緊急度', '中身設計コード', '容量_L'], ascending=[True, True, True, True, False]).copy()

                    # 🌟 確定済み入庫予定のライン・品目情報付与
                    confirmed_jobs = []
                    if use_confirmed and consider_iko and iko_entries:
                        line_info_dict = df_final.drop_duplicates(subset=['品目コード']).set_index('品目コード')[['製造ライン', '品目名', '容量_L', '中身設計コード']].to_dict('index')
                        for entry in iko_entries:
                            i_code = entry['品目コード']
                            i_qty = entry['数量']
                            i_date = entry['日付']
                            if i_qty <= 0:
                                continue
                            if i_code in line_info_dict:
                                info = line_info_dict[i_code]
                                i_line = info['製造ライン']
                                i_name = info['品目名']
                                i_vol = info['容量_L']
                                i_recipe = info['中身設計コード']
                            else:
                                zai_row_lookup = df_zai_in_zai[df_zai_in_zai['品目コード'] == i_code]
                                i_name = zai_row_lookup['品目名'].iloc[0] if not zai_row_lookup.empty else i_code
                                i_vol = extract_volume_safe(i_name)
                                i_recipe = extract_content_code(i_code)
                                if factory_mode == "本社":
                                    # 本社ライン割り当てルールと同一ロジック
                                    if i_code in ('H0690020', 'H0690000', 'H0690030', 'H0390000'):
                                        i_line = '6号機'
                                    elif i_code in HONSHA_FIXED_LINE:
                                        i_line = HONSHA_FIXED_LINE[i_code]
                                    elif i_code == 'H0620030' or any(k in i_name for k in ['再生材', 'もう一土元気']) or any(k in i_name for k in ['腐葉土', '堆肥', '特大袋']):
                                        i_line = '3号機'
                                    elif i_vol < 10:
                                        i_line = 'その他'
                                    elif i_vol <= 25:
                                        i_line = '5号機'
                                    elif i_vol <= 26:
                                        i_line = '2号機'
                                    else:
                                        i_line = '6号機'
                                else:
                                    i_line = 'その他'
                            confirmed_jobs.append({
                                '日付': i_date, 'ライン': i_line, '配合コード': i_recipe,
                                '品目コード': i_code, '品目名': i_name, '数量': i_qty, '容量_L': i_vol
                            })

                    del df_master_combined, df_final, grouped
                    gc.collect()

                    # =====================================================================
                    # 🌟 lines_list（本社：その他を追加）
                    # =====================================================================
                    lines_list = ["1号機", "2号機", "3号機", "4号機", "5号機", "6号機", "その他"] if factory_mode == "関西工場" else ["2号機", "3号機", "5号機", "6号機", "その他"]
                    def _build_line_queue(df_line):
                        """月度順にグループ化し、各月内で容量近接ソートを適用"""
                        jobs_seq = []
                        for _m in sorted(df_line['対象月度'].unique(), key=lambda x: _month_order.get(x, 99)):
                            jobs_seq.extend(sort_jobs_by_size_proximity(df_line[df_line['対象月度'] == _m]))
                        return jobs_seq
                    queues_base = {line: _build_line_queue(df_final_sorted[df_final_sorted['製造ライン'] == line]) for line in lines_list}

                    def run_sim(ov_mins):
                        queues = copy.deepcopy(queues_base)
                        cur_idx = {l: 0 for l in queues}
                        for l in queues:
                            for j in queues[l]: j['rem'] = j['計画製造袋数']

                        loop_d = get_next_w_date(period_start, holidays_input)
                        day_cnt = 1; sched = []

                        while True:
                            active = [l for l in lines_list if cur_idx[l] < len(queues[l]) and queues[l][cur_idx[l]]['rem'] > 0 or any(j['rem'] > 0 and job_can_support(l, j, factory_mode) for ol in lines_list if ol != l for j in queues[ol])]
                            if not active: break

                            if factory_mode == "関西工場":
                                run_today = active
                            else:
                                _fa = [l for l in ['2号機', '3号機', '5号機', '6号機'] if l in active]
                                _oa = ['その他'] if 'その他' in active else []
                                if len(_fa) == 4:
                                    run_today = _fa + _oa
                                elif any(l in _fa for l in ['3号機', '5号機']) and ('5号機' in _fa or not any(l in _fa for l in ['2号機', '6号機'])):
                                    run_today = [l for l in ['3号機', '5号機'] if l in _fa] + _oa
                                else:
                                    run_today = [l for l in ['2号機', '6号機'] if l in _fa] + _oa

                            # 稼働時間：月〜木430分/金400分 を基本とし、月によって調整
                            # 7月・8月・12月: 1時間短縮、3月・4月: 1時間延長（カレンダー月基準）
                            _base_cap = 400.0 if loop_d.weekday() == 4 else 430.0
                            _month_adj = -60.0 if loop_d.month in (7, 8, 12) else (60.0 if loop_d.month in (3, 4) else 0.0)
                            cap_limit = _base_cap + _month_adj + ov_mins
                            w_kanji = ["月", "火", "水", "木", "金", "土", "日"][loop_d.weekday()]
                            d_str_disp = loop_d.strftime("%Y/%m/%d")

                            for line in run_today:
                                spent = 0.0; p_rec = None; p_vol = None

                                # 特殊清掃が必要な品目は、その日まだ何も製造していない（＝機械が清掃済みの状態の）
                                # タイミングでのみ繰り上げて先頭に持ってくる。前日から継続中のジョブは中断しない。
                                _cur = cur_idx[line]
                                if _cur < len(queues[line]) and queues[line][_cur]['rem'] == queues[line][_cur]['計画製造袋数']:
                                    for _scan_idx in range(_cur, len(queues[line])):
                                        _cand = queues[line][_scan_idx]
                                        if _cand['rem'] <= 0:
                                            continue
                                        if _month_order.get(_cand.get('対象月度', _first_month), 0) > _month_order.get(loop_d.month, 99):
                                            break
                                        if any(k in str(_cand.get('品目名', '')) for k in SPECIAL_CLEANING_KEYWORDS):
                                            if _scan_idx != _cur:
                                                queues[line].insert(_cur, queues[line].pop(_scan_idx))
                                            break

                                day_confirmed = [cj for cj in confirmed_jobs if cj['日付'] == loop_d and cj['ライン'] == line]
                                for cj in day_confirmed:
                                    sp_min_c = get_sp(line, cj['容量_L'], factory_mode, cj['品目コード']) / 60
                                    dur_c = cj['数量'] / sp_min_c if sp_min_c > 0 else 0.0
                                    sw_c = 5.0 if spent > 0 and p_rec == cj['配合コード'] and p_vol and p_vol > cj['容量_L'] else (10.0 if spent > 0 else 0.0)
                                    sched.append({'稼働日': f"{day_offset + day_cnt}日目", '製造日': d_str_disp, '曜日': w_kanji, '製造ライン': line, '配合コード': cj['配合コード'], '品目コード': cj['品目コード'], '品目名': cj['品目名'], '指示数量(袋)': int(cj['数量']), '製造時間(分)': round(dur_c, 1), '切り替え(分)': round(sw_c, 1), '合計拘束時間(分)': round(sw_c + dur_c, 1), '備考': '製造指示済', '製造理由': '製造指示済', '対象月度': f"{cj['日付'].month}月度", 't_start': spent + sw_c, 't_end': spent + sw_c + dur_c})
                                    spent += sw_c + dur_c
                                    p_rec = cj['配合コード']; p_vol = cj['容量_L']

                                while spent < cap_limit:
                                    idx = cur_idx[line]
                                    if idx < len(queues[line]):
                                        job = queues[line][idx]
                                        # 同月度内計画: 対象月度がまだ来ていないジョブは当日処理しない
                                        if _month_order.get(job.get('対象月度', _first_month), 0) > _month_order.get(loop_d.month, 99):
                                            break
                                        sw = 5.0 if spent > 0 and p_rec == job['中身設計コード'] and p_vol and p_vol > job['容量_L'] else (10.0 if spent > 0 else 0.0)
                                        avail = cap_limit - spent - sw
                                        if avail <= 5.0: break

                                        sp_min = get_sp(line, job['容量_L'], factory_mode, job['品目コード']) / 60
                                        max_b = avail * sp_min

                                        if job['rem'] <= max_b:
                                            b_make = job['rem']; dur = b_make / sp_min
                                            sched.append({'稼働日': f"{day_offset + day_cnt}日目", '製造日': d_str_disp, '曜日': w_kanji, '製造ライン': line, '配合コード': job['中身設計コード'], '品目コード': job['品目コード'], '品目名': job['品目名'], '指示数量(袋)': int(b_make), '製造時間(分)': round(dur, 1), '切り替え(分)': round(sw, 1), '合計拘束時間(分)': round(sw + dur, 1), '備考': '全量完了', '製造理由': job['製造理由'], '対象月度': f"{job.get('対象月度', '')}月度", 't_start': spent + sw, 't_end': spent + sw + dur})
                                            spent += sw + dur; job['rem'] = 0; cur_idx[line] += 1
                                            p_rec = job['中身設計コード']; p_vol = job['容量_L']
                                        else:
                                            b_make = math.floor(max_b)
                                            if b_make <= 0: break
                                            dur = b_make / sp_min
                                            sched.append({'稼働日': f"{day_offset + day_cnt}日目", '製造日': d_str_disp, '曜日': w_kanji, '製造ライン': line, '配合コード': job['中身設計コード'], '品目コード': job['品目コード'], '品目名': job['品目名'], '指示数量(袋)': int(b_make), '製造時間(分)': round(dur, 1), '切り替え(分)': round(sw, 1), '合計拘束時間(分)': round(sw + dur, 1), '備考': '翌日へ継続', '製造理由': job['製造理由'], '対象月度': f"{job.get('対象月度', '')}月度", 't_start': spent + sw, 't_end': spent + sw + dur})
                                            spent += sw + dur; job['rem'] -= b_make
                                            p_rec = job['中身設計コード']; p_vol = job['容量_L']; break
                                    else:
                                        sup_found = False
                                        for o_line in lines_list:
                                            if o_line == line: continue
                                            for job in queues[o_line]:
                                                if job['rem'] > 0 and _month_order.get(job.get('対象月度', _first_month), 0) <= _month_order.get(loop_d.month, 99) and job_can_support(line, job, factory_mode):
                                                    sw = 10.0; avail = cap_limit - spent - sw
                                                    if avail <= 5.0: break
                                                    sp_min = (646 if line == '5号機' else 500) / 60 if factory_mode == "関西工場" else ((730 if job['容量_L'] in [12, 14] else 650) if line == '5号機' else 400) / 60
                                                    max_b = avail * sp_min
                                                    if job['rem'] <= max_b:
                                                        b_make = job['rem']; dur = b_make / sp_min
                                                        sched.append({'稼働日': f"{day_offset + day_cnt}日目", '製造日': d_str_disp, '曜日': w_kanji, '製造ライン': line, '配合コード': job['中身設計コード'], '品目コード': job['品目コード'], '品目名': job['品目名'], '指示数量(袋)': int(b_make), '製造時間(分)': round(dur, 1), '切り替え(分)': round(sw, 1), '合計拘束時間(分)': round(sw + dur, 1), '備考': f"★{o_line}応援(全量)", '製造理由': job['製造理由'], '対象月度': f"{job.get('対象月度', '')}月度", 't_start': spent + sw, 't_end': spent + sw + dur})
                                                        spent += sw + dur; job['rem'] = 0; sup_found = True; p_rec = job['中身設計コード']; p_vol = job['容量_L']; break
                                                    else:
                                                        b_make = math.floor(max_b)
                                                        if b_make <= 0: break
                                                        dur = b_make / sp_min
                                                        sched.append({'稼働日': f"{day_offset + day_cnt}日目", '製造日': d_str_disp, '曜日': w_kanji, '製造ライン': line, '配合コード': job['中身設計コード'], '品目コード': job['品目コード'], '品目名': job['品目名'], '指示数量(袋)': int(b_make), '製造時間(分)': round(dur, 1), '切り替え(分)': round(sw, 1), '合計拘束時間(分)': round(sw + dur, 1), '備考': f"★{o_line}応援(継続)", '製造理由': job['製造理由'], '対象月度': f"{job.get('対象月度', '')}月度", 't_start': spent + sw, 't_end': spent + sw + dur})
                                                        spent += sw + dur; job['rem'] -= b_make; sup_found = True; p_rec = job['中身設計コード']; p_vol = job['容量_L']; break
                                            if sup_found: break
                                        if not sup_found: break
                            loop_d = get_next_w_date(loop_d + datetime.timedelta(days=1), holidays_input)
                            day_cnt += 1
                            if day_cnt > 60: break
                        return sched, day_cnt - 1

                    ov_res = 0
                    full_sched, gen_days = run_sim(0)
                    if gen_days > target_bd:
                        _fit_found = False
                        _last_ts, _last_td, _last_ov = full_sched, gen_days, 0
                        for t_ov in [30, 60, 90, 120, 150, 180, 210]:
                            ts, td = run_sim(t_ov)
                            _last_ts, _last_td, _last_ov = ts, td, t_ov
                            if td <= target_bd:
                                ov_res = t_ov; full_sched = ts; gen_days = td; _fit_found = True; break
                        if not _fit_found:
                            # 最大残業でも月内に収まらない場合は、最も日数の短い（最大OT）結果を採用
                            ov_res = _last_ov; full_sched = _last_ts; gen_days = _last_td

                    df_final_sorted['対象月度'] = month_label
                    for _j in full_sched: _j['対象月度'] = month_label
                    return full_sched, df_final_sorted, gen_days, ov_res

                # =====================================================================
                # 🌟 月度ごとに独立して計画を実行（各月の計画はその月内でスケジュール）
                # =====================================================================
                def _year_for_month(m):
                    _t = datetime.date.today()
                    y = _t.year
                    if m < _t.month - 6: y += 1
                    return y

                full_sched = []
                _df_frames = []
                _day_offset = 0
                _cursor = start_date
                _summary = []

                for _pi, (_m_num, _p_idx_d, _a_idx_d) in enumerate(month_col_pairs):
                    # この月度の計画残数をdf_m_distinctの月別列から取得
                    _zan_col = f'残数_{_m_num}'
                    if _zan_col in df_m_distinct.columns:
                        _df_mp = df_m_distinct[['品目コード', '品目名_計画書']].copy()
                        _df_mp['選択月_計画残数'] = df_m_distinct[_zan_col].values
                    else:
                        _df_mp = df_m_distinct[['品目コード', '品目名_計画書', '選択月_計画残数']].copy()

                    # 期間開始日・終了日を決定
                    if _pi == 0:
                        _p_start = start_date
                    else:
                        _m_first = datetime.date(_year_for_month(_m_num), _m_num, 1)
                        _p_start = max(_m_first, _cursor)
                    _p_end = _get_month_end(f"{_m_num}月", datetime.date.today())
                    if plan_to_yearend:
                        _target_bd = _count_business_days(_p_start, _p_end, holidays_input)
                    else:
                        _target_bd = target_days

                    _sched_p, _dfs_p, _days_p, _ov_p = _plan_period(
                        _df_mp,
                        use_inventory=(_pi == 0),
                        use_confirmed=(_pi == 0),
                        period_start=_p_start,
                        target_bd=_target_bd,
                        day_offset=_day_offset,
                        month_label=_m_num
                    )
                    full_sched.extend(_sched_p)
                    if not _dfs_p.empty:
                        _df_frames.append(_dfs_p)
                    _day_offset += _days_p
                    if _sched_p:
                        _last_d = max(datetime.datetime.strptime(j['製造日'], "%Y/%m/%d").date() for j in _sched_p)
                        _cursor = _last_d + datetime.timedelta(days=1)
                    else:
                        _cursor = max(_cursor, _p_start)
                    _summary.append((_m_num, _days_p, _ov_p, _p_end, _target_bd))

                df_final_sorted = pd.concat(_df_frames, ignore_index=True) if _df_frames else pd.DataFrame()

                # =====================================================================
                # 🌟 天川月間製造計画書の反映（本社モードのみ）
                #   MRP連携用にシート1（バッチ集計）へ数量を反映する。
                #   タイムテーブル（シート2・3）は本社ライン用のため対象外。
                # =====================================================================
                if factory_mode == "本社":
                    try:
                        safe_seek(file_gekkan)
                        _xl_tk = pd.ExcelFile(file_gekkan)
                        _tk_sheets = [s for s in _xl_tk.sheet_names if '天川' in s and '計画' in s]
                        if _tk_sheets:
                            _df_tk = pd.read_excel(_xl_tk, sheet_name=_tk_sheets[0], header=None)
                            # 見出し行（商品CD）の検出
                            _tk_item_row = next((i for i in range(min(15, len(_df_tk))) if any(kw in [str(v).strip() for v in _df_tk.iloc[i].values] for kw in ['商品CD', '商品コード', '品目コード', '商品CODE'])), 2)
                            # 月ラベルスキャン
                            _tk_labels = []
                            _tk_seen = set()
                            for _r in range(0, _tk_item_row + 1):
                                for _c in range(_df_tk.shape[1]):
                                    if _c in _tk_seen: continue
                                    _v = _df_tk.iloc[_r, _c]
                                    if _v is not None:
                                        _s = str(_v).strip()
                                        _m = re.search(r'(\d{1,2})\s*月', _s)
                                        if _m and not re.search(r'\d{1,2}\s*月\s*\d{1,2}\s*日', _s):
                                            _mn = int(_m.group(1))
                                            if 1 <= _mn <= 12:
                                                _tk_labels.append((_c, _mn)); _tk_seen.add(_c)
                            _tk_labels.sort(key=lambda x: x[0])
                            _tk_dedup = []
                            for _cp, _mn in _tk_labels:
                                if _tk_dedup and _tk_dedup[-1][1] == _mn: continue
                                _tk_dedup.append((_cp, _mn))
                            _tk_labels = _tk_dedup
                            # 見出し行（自体と1つ下）の値
                            _tk_hdr_rows = [_tk_item_row] + ([_tk_item_row + 1] if _tk_item_row + 1 < len(_df_tk) else [])
                            _tk_hdr_vals = {r: [str(v).strip() for v in _df_tk.iloc[r].values] for r in _tk_hdr_rows}
                            # ロット列
                            _tk_lot_col = None
                            for _hr in _tk_hdr_rows:
                                _tk_lot_col = next((c for c, t in enumerate(_tk_hdr_vals[_hr]) if 'ロット' in t or 'ﾛｯﾄ' in t), None)
                                if _tk_lot_col is not None: break

                            def _tk_find_block(mn):
                                for i, (cp, m) in enumerate(_tk_labels):
                                    if m == mn:
                                        return cp, (_tk_labels[i+1][0] if i+1 < len(_tk_labels) else _df_tk.shape[1])
                                return None, None

                            def _tk_find_pa(bs, be):
                                if bs is None: return None, None
                                rng = list(range(bs, be))
                                for _hr in _tk_hdr_rows:
                                    vals = _tk_hdr_vals[_hr]
                                    p = next((c for c in rng if c < len(vals) and '製造予定' in vals[c]), None)
                                    a = next((c for c in rng if c < len(vals) and '製造実績' in vals[c]), None)
                                    if p is not None and a is not None: return p, a
                                return None, None

                            _tk_data = _df_tk.iloc[_tk_item_row + 1:]
                            _tk_rows = []
                            for _m_num_tk in months_to_plan:
                                _bs, _be = _tk_find_block(_m_num_tk)
                                _p_c, _a_c = _tk_find_pa(_bs, _be)
                                if _p_c is None: continue
                                for _ri in range(len(_tk_data)):
                                    _code_tk = str(_tk_data.iloc[_ri, 0]).strip()
                                    if not re.match(r'^[A-Za-z]\d', _code_tk): continue
                                    _pv = pd.to_numeric(_tk_data.iloc[_ri, _p_c], errors='coerce')
                                    _av = pd.to_numeric(_tk_data.iloc[_ri, _a_c], errors='coerce')
                                    _pv = 0 if pd.isna(_pv) else _pv
                                    _av = 0 if pd.isna(_av) else _av
                                    _zan_tk = max(0, _pv - _av)
                                    if _zan_tk <= 0: continue
                                    _name_tk = str(_tk_data.iloc[_ri, 1]).strip()
                                    _lot_tk = 0.0
                                    if _tk_lot_col is not None:
                                        _lv_tk = pd.to_numeric(_tk_data.iloc[_ri, _tk_lot_col], errors='coerce')
                                        if not pd.isna(_lv_tk) and _lv_tk > 0: _lot_tk = float(_lv_tk)
                                    _bags_tk = int(math.ceil(_zan_tk / _lot_tk) * _lot_tk) if _lot_tk > 0 else int(_zan_tk)
                                    _vol_tk = extract_volume_safe(_name_tk)
                                    if is_kg_product(_name_tk):
                                        _m3_tk = round(_bags_tk * get_kg_weight(_name_tk), 1)
                                    else:
                                        _m3_tk = round(_bags_tk * max(_vol_tk, 1) / 900.0, 1)
                                    _tk_rows.append({
                                        '対象月度': _m_num_tk, '品目コード': _code_tk, '品目名': _name_tk,
                                        '製造ライン': '天川', '中身設計コード': extract_content_code(_code_tk),
                                        '現在の在庫': np.nan, '安全在庫数': np.nan, '安全割れ不足数': 0.0,
                                        '今月の計画残数': _zan_tk, '生産ロット': _lot_tk,
                                        '製造決定_m3': _m3_tk, '計画製造袋数': _bags_tk,
                                        '製造理由': '天川計画', '容量_L': _vol_tk,
                                    })
                            if _tk_rows:
                                df_final_sorted = pd.concat([df_final_sorted, pd.DataFrame(_tk_rows)], ignore_index=True)
                                st.info(f"⛰️ 天川月間製造計画書（{_tk_sheets[0]}）: {len(_tk_rows)}件（品目×月度）をバッチ集計シートへ反映しました。タイムテーブルは対象外です。")
                    except Exception as _e_tk:
                        st.warning(f"⚠️ 天川計画書の読込に失敗しました: {_e_tk}")

                if df_final_sorted.empty:
                    st.warning("計画対象となる品目がありません。")
                    st.stop()

                # 月度別サマリー表示
                for (_m_num, _days_p, _ov_p, _p_end, _tbd_p) in _summary:
                    if _days_p == 0:
                        st.info(f"ℹ️ {_m_num}月度: 計画対象なし")
                    elif _days_p > _tbd_p:
                        st.error(f"🚨 {_m_num}月度: 最大残業（{_ov_p}分/日）でも月内（{_p_end.strftime('%m/%d')}・{_tbd_p}営業日）に収まらず、【{_days_p}日間】かかり翌月度に食い込みます。翌月度の計画開始はその分後ろ倒しになります。")
                    elif _ov_p > 0:
                        st.warning(f"📢 {_m_num}月度: 月内（{_p_end.strftime('%m/%d')}まで）に収めるため毎日一律【{_ov_p}分】の残業が必要です（{_days_p}日間）。")
                    else:
                        st.success(f"🟢 {_m_num}月度: 残業不要・【{_days_p}日間】で完了します。")

                lines_list = ["1号機", "2号機", "3号機", "4号機", "5号機", "6号機", "その他"] if factory_mode == "関西工場" else ["2号機", "3号機", "5号機", "6号機", "その他"]

                wb = Workbook(); wb.remove(wb.active)
                navy = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                w_font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
                r_font = Font(name="Meiryo UI", size=10); b_font = Font(name="Meiryo UI", size=10, bold=True)
                b_all = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))

                ws1 = wb.create_sheet(title="製造品目・バッチ集計"); ws1.views.sheetView[0].showGridLines = True
                ws1.append(["対象月度", "品目コード", "品目名", "製造ライン", "配合レシピ", "現在の在庫", "安全在庫数", "安全割れ不足数", "当月度の計画残数", "生産ロット", "決定製造m3", "最終製造総数(袋)", "製造理由"])
                for _, r in df_final_sorted.iterrows(): ws1.append([f"{r.get('対象月度', '')}月度", r['品目コード'], r['品目名'], r['製造ライン'], r['中身設計コード'], r['現在の在庫'], r['安全在庫数'], r['安全割れ不足数'], r['今月の計画残数'], (int(r['生産ロット']) if r.get('生産ロット', 0) > 0 else ""), r['製造決定_m3'], r['計画製造袋数'], r['製造理由']])

                ws2 = wb.create_sheet(title="日別・号機別製造計画"); ws2.views.sheetView[0].showGridLines = True
                ws2.append(["稼働日", "製造日", "曜日", "対象月度", "製造ライン", "配合コード", "品目コード", "品目名", "指示数量(袋)", "製造時間(分)", "切り替え(分)", "合計拘束時間(分)", "備考", "製造理由"])
                for j in full_sched: ws2.append([j['稼働日'], j['製造日'], j['曜日'], f"{j.get('対象月度', '')}月度" if j.get('対象月度', '') != '' else '', j['製造ライン'], j['配合コード'], j['品目コード'], j['品目名'], j['指示数量(袋)'], j['製造時間(分)'], j['切り替え(分)'], j['合計拘束時間(分)'], j['備考'], j['製造理由']])

                ws3 = wb.create_sheet(title="日別・30分タイムテーブル"); ws3.views.sheetView[0].showGridLines = True
                slots = ["8:00〜8:30", "8:30〜9:00", "9:00〜9:30", "9:30〜10:00", "10:00〜10:10(休憩)", "10:10〜10:30", "10:30〜11:00", "11:00〜11:30", "11:30〜12:00", "12:00〜13:00(昼休)", "13:00〜13:30", "13:30〜14:00", "14:00〜14:30", "14:30〜15:00", "15:00〜15:10(休憩)", "15:10〜15:30", "15:30〜16:00", "16:00〜16:30", "16:30〜17:00", "17:00〜17:30", "17:30〜18:00", "18:00〜18:30", "18:30〜19:00", "19:00〜19:30", "19:30〜20:00"]
                ws3.append(["稼働日", "製造日", "ライン"] + slots)

                u_days = []
                seen_d = set()
                for j in full_sched:
                    k = (j['稼働日'], j['製造日'])
                    if k not in seen_d: seen_d.add(k); u_days.append(k)

                mat_map = {}
                for (ds, ddt) in u_days:
                    wk = ["月", "火", "水", "木", "金", "土", "日"][datetime.datetime.strptime(ddt, "%Y/%m/%d").weekday()]
                    for ln in lines_list:
                        ldisp = {"1号機": "NO.1", "2号機": "NO.2", "3号機": "NO.3", "4号機": "NO.4", "5号機": "NO.5", "6号機": "NO.6"}.get(ln, ln)
                        row_arr = [ds, f"{ddt} ({wk})", ldisp] + [""] * 25
                        ws3.append(row_arr)
                        mat_map[(ds, ln)] = ws3[ws3.max_row]

                s_rng = {0:(0,30), 1:(30,60), 2:(60,90), 3:(90,120), 4:(None,"休"), 5:(120,140), 6:(140,170), 7:(170,200), 8:(200,230), 9:(None,"昼"), 10:(230,260), 11:(260,290), 12:(290,320), 13:(320,350), 14:(None,"休"), 15:(350,370), 16:(370,400), 17:(400,430), 18:(430,460), 19:(460,490), 20:(490,520), 21:(520,550), 22:(550,580), 23:(580,610), 24:(610,640)}

                REASON_FILLS = {
                    '現在庫がマイナス': PatternFill(start_color="F8C9C4", end_color="F8C9C4", fill_type="solid"),
                    '安全在庫割れ':     PatternFill(start_color="FCEAA4", end_color="FCEAA4", fill_type="solid"),
                    '計画未達':         PatternFill(start_color="C9E4F8", end_color="C9E4F8", fill_type="solid"),
                    '製造指示済':       PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
                }
                REASON_PRIORITY = {'現在庫がマイナス': 4, '安全在庫割れ': 3, '計画未達': 2, '製造指示済': 1}
                cell_reason_priority = {}

                for j in full_sched:
                    t_cell_row = mat_map.get((j['稼働日'], j['製造ライン']))
                    if t_cell_row:
                        sm = j['t_start']; em = j['t_end']
                        for si in range(25):
                            if si in [4, 9, 14]:
                                t_cell_row[si+3].value = "小休憩" if si!=9 else "昼休憩"; continue
                            st_s, ed_s = s_rng[si]
                            if max(sm, st_s) < min(em, ed_s) - 1e-5:
                                pfx = "★(応援)\n" if "応援" in j['備考'] else ""
                                cur_v = t_cell_row[si+3].value or ""
                                add_t = f"{pfx}{j['品目名']}\n({j['指示数量(袋)']}袋)"
                                t_cell_row[si+3].value = f"{cur_v}＋\n{add_t}" if cur_v else add_t

                                cell_key = (t_cell_row[si+3].row, t_cell_row[si+3].column)
                                this_priority = REASON_PRIORITY.get(j.get('製造理由', ''), 0)
                                if this_priority > 0 and this_priority >= cell_reason_priority.get(cell_key, 0):
                                    cell_reason_priority[cell_key] = this_priority
                                    fill = REASON_FILLS.get(j['製造理由'])
                                    if fill:
                                        t_cell_row[si+3].fill = fill

                for sheet in [ws1, ws2, ws3]:
                    sheet.row_dimensions[1].height = 26
                    for c in sheet[1]: c.fill = navy; c.font = w_font; c.alignment = Alignment(horizontal="center", vertical="center")
                    for r_idx in range(2, sheet.max_row+1):
                        sheet.row_dimensions[r_idx].height = 20 if sheet!=ws3 else 60
                        is_z = (r_idx%2 == 0) if sheet!=ws3 else ((r_idx-2)//len(lines_list)%2 == 0)
                        for c in sheet[r_idx]:
                            c.font = r_font; c.border = b_all
                            if sheet==ws3:
                                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                already_colored = (c.row, c.column) in cell_reason_priority
                                if c.column in [2,3]: c.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                                elif c.column in [8,13,18]: c.fill = PatternFill(start_color="E4DFEC" if c.column==13 else "EAEAEA", end_color="E4DFEC" if c.column==13 else "EAEAEA", fill_type="solid"); c.font = b_font
                                elif c.column>3 and is_z and not already_colored: c.fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
                            else:
                                if is_z: c.fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
                                c.alignment = Alignment(horizontal="right" if isinstance(c.value, (int,float)) else "left", vertical="center")
                                if isinstance(c.value, (int,float)): c.number_format = "#,##0"
                    if sheet!=ws3:
                        for col in sheet.columns: sheet.column_dimensions[get_column_letter(col[0].column)].width = max(max(sum(2 if ord(char)>128 else 1 for char in str(cell.value or '')) for cell in col)+3, 12)
                        sheet.freeze_panes = "A2"
                    else:
                        nl = len(lines_list)
                        for di in range(len(u_days)):
                            sheet.merge_cells(start_row=2+di*nl, start_column=1, end_row=2+di*nl+nl-1, end_column=1)
                            sheet.merge_cells(start_row=2+di*nl, start_column=2, end_row=2+di*nl+nl-1, end_column=2)
                        sheet.column_dimensions['A'].width=12; sheet.column_dimensions['B'].width=16; sheet.column_dimensions['C'].width=12
                        for ci in range(4, sheet.max_column+1): sheet.column_dimensions[get_column_letter(ci)].width = 23
                        sheet.freeze_panes = "D2"

                out_io = io.BytesIO(); wb.save(out_io); out_io.seek(0)
                _fname_period = f"{target_month}度〜10月度(期末)" if plan_to_yearend else f"{target_month}度"
                st.download_button("📊 指示スケジュール表(.xlsx)をダウンロード", out_io, f"【確定版】{factory_mode}_{_fname_period}_スケジュール表.xlsx")
            except Exception as e: st.error(f"計算実行エラー: {e}")
